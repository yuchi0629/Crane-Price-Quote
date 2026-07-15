import base64
import json
import os
import re
import sys
from io import BytesIO
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from tkinter import BOTH, LEFT, RIGHT, VERTICAL, BooleanVar, Canvas, StringVar, Text, Tk, Toplevel, filedialog, messagebox
from tkinter import ttk

from PIL import Image as PILImage, ImageTk
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from copy import copy
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import BaseDocTemplate, Frame, Image, PageTemplate, Paragraph, Table, TableStyle


BUNDLE_DIR = Path(getattr(sys, "_MEIPASS", Path(__file__).resolve().parent))
APP_DIR = Path(sys.executable).resolve().parent if getattr(sys, "frozen", False) else Path(__file__).resolve().parent
if getattr(sys, "frozen", False):
    USER_DATA_ROOT = Path(os.environ.get("APPDATA", APP_DIR)) / "ZoomlionTowerCraneQuotation"
    DATA_DIR = USER_DATA_ROOT / "data"
    OUTPUT_DIR = Path(os.environ.get("USERPROFILE", APP_DIR)) / "Documents" / "中联塔机报价单输出"
else:
    DATA_DIR = APP_DIR / "data"
    OUTPUT_DIR = APP_DIR / "output" / "pdf"
DB_FILE = DATA_DIR / "quotation_database.json"
SETTINGS_FILE = DATA_DIR / "user_settings.json"
INITIAL_DB_FILE = BUNDLE_DIR / "data" / "quotation_database.json"
WINDOW_ICON_FILE = BUNDLE_DIR / "assets" / "zoomlion.ico"
HEADER_ICON_FILE = BUNDLE_DIR / "assets" / "crane.png"
APP_TITLE = "中联塔机配置确认及报价单生成软件 V1.0"
CONFIG_DATA_RESET_VERSION = 2
EMBEDDED_LOGO_PNG = (
    "iVBORw0KGgoAAAANSUhEUgAAAGwAAAAfCAYAAAAC0CiiAAAAAXNSR0IArs4c6QAAAARnQU1BAACxjwv8"
    "YQUAAAAJcEhZcwAADsQAAA7EAZUrDhsAAANKSURBVGhD7ZoxSBtRGMf/7WQLBUsKUWgiImZwEMngECxdQ"
    "gfXFIcOJUNxkkxO2k6tTplCJ6dMDtIM7eBQsoiSoUMQCw4RkZpCFSoVHOrWvvfue5d3d8kl7y6lPHg/CP"
    "fd3efly/2/933fK72z08z9gcUY7tLRYghWMMOwghmGFcwwrGCGYQUzDCuYYVjBDMMKZhh6/9JxcIql0iW"
    "d9KAwg531UTphdPmbzGoW716M0JmPgfxvsVtsonpEp0hirTmFOTpz8PnMTqJSHceYet0fK3G40cBmjRnd7"
    "g/4e9xnMPKVHJYXHFsS+h0hDHmFPUDxpfxy/mIaXQVulZtYyp7ikM4ddP1VLvHlgExJ+xcarqDDIHp89VJ"
    "Y7HroCbYwBbYifZ8Z5Ol2ZnUaiynHvtg+cbObZ6Dr/3ESGXH1Epsb18Li6PpLMrMPxLG+5713sf8TLXaU9"
    "+MSNT6HsHt6xFxhPOuOUecmW9qdsnCNT+Ubx/RcZ6TGsbJKL7HWxm6bG7r+CtP3nRdWu1Ky+BbNz/x5SaS"
    "n6bmxiBFfIekkdO0YW/4qEIFYgnWyjvUQT9+6ckT0lMgOY08eUVbe4Ps3dtD1V5lMIDfLDaUsynJYSGDeu"
    "RKPOPEhgeVKUlj10tegoJpEF4w13xJlXb7ib/h9SN1Dmszz81uyQgj1H0H2mbcsynKYfxp8uf+Efr+HtZK"
    "1AjduUH3zAxfiYjQiCnaNLWq+vJ77J6C+tH/jnMx02jtddaWPv5vhoiR1yuG8blxRGeD3zK1Trz86w/vtA"
    "ZK0BxEEU/oWG5VXfOOsYIL6Csuoxn4wOLkCeIl5PMEOuv5+Ug+pLLKStN8ph1qrPoy48QlG3dLYKp/gw4k"
    "wtdEWzNO3xL6mC6wRPxclwBl5X6sZpZRSFFLOVKnrH2AEi6+oT5TPhl8OY8dHKKWxRROnLnqCqcHxUTXL9"
    "iX+D42vc+tZFEXWy30K3Zf7GL6RVQYVXf8ACwl3ezFwOWSTm/s9/FPs3V9ix0eoz4mClmCHe8FNY29Y1ld"
    "zqMiRV0HsYwKrU9ffzyjmaRUMtRy6xI1Pwp7zVu7d9LH/CccwIk6Jlv+FFcwwrGCGYQUzDCuYYVjBDMMKZ"
    "hhWMMOwghkF8Bd9LbQQl4zAawAAAABJRU5ErkJggg=="
)

COMPONENT_KEYWORDS = [
    ("起重臂", "Jib"),
    ("平衡臂", "Counter-jib"),
    ("司机室", "Cabin"),
    ("上支座", "Upper support"),
    ("回转支承", "Slewing bearing"),
    ("下支座", "Lower support"),
    ("吊钩", "Hook"),
    ("标准节", "Mast section"),
    ("底架", "Base frame"),
    ("支腿组件", "Support leg assembly"),
    ("基础预埋", "Support leg assembly", "支腿组件"),
    ("机构", "Mechanism"),
    ("钢丝绳", "Wire rope"),
    ("电控总成", "Electric control assembly"),
]

LANG_OPTIONS = {"中文": "zh", "English": "en", "Français": "fr", "Deutsch": "de"}

TRANSLATIONS = {
    "中联塔机报价单": {"en": "TOWER CRANE QUOTATION", "fr": "DEVIS POUR GRUE A TOUR", "de": "ANGEBOT FUER TURMDREHKRAN"},
    "报价日期": {"en": "Date", "fr": "Date", "de": "Datum"},
    "客户名称": {"en": "Customer", "fr": "Client", "de": "Kunde"},
    "报价单位": {"en": "Quotation unit", "fr": "Unite emettrice", "de": "Angebotseinheit"},
    "报价人": {"en": "Quoted by", "fr": "Emetteur", "de": "Ansprechpartner"},
    "联系电话": {"en": "Phone", "fr": "Telephone", "de": "Telefon"},
    "联系邮箱": {"en": "Email", "fr": "E-mail", "de": "E-Mail"},
    "基本参数": {"en": "BASIC PARAMETER", "fr": "PARAMETRES DE BASE", "de": "GRUNDPARAMETER"},
    "产品型号": {"en": "Model", "fr": "Modele", "de": "Modell"},
    "塔机类型": {"en": "Type", "fr": "Type", "de": "Typ"},
    "独立高度": {"en": "HUH", "fr": "Hauteur autoportante (HUH)", "de": "Freistehende Hoehe (HUH)"},
    "最大臂长": {"en": "Max jib length", "fr": "Portee maximale", "de": "Max. Auslegerlaenge"},
    "最大起重量": {"en": "Max load", "fr": "Charge maximale", "de": "Max. Traglast"},
    "安装形式": {"en": "Installation form", "fr": "Type d'installation", "de": "Aufstellungsart"},
    "容绳量": {"en": "Rope capacity", "fr": "Capacite de cable", "de": "Seilkapazitaet"},
    "塔身种类": {"en": "Mast type", "fr": "Type de mat", "de": "Masttyp"},
    "价格": {"en": "Price", "fr": "Prix", "de": "Preis"},
    "单价": {"en": "UNIT PRICE", "fr": "PRIX UNITAIRE", "de": "EINHEITSPREIS"},
    "主要部件配置": {"en": "MAIN COMPONENT CONFIGURATION", "fr": "CONFIGURATION DES COMPOSANTS PRINCIPAUX", "de": "HAUPTKOMPONENTENKONFIGURATION"},
    "部件名称": {"en": "Component", "fr": "Composant", "de": "Komponente"},
    "代号": {"en": "Code", "fr": "Code", "de": "Code"},
    "数量": {"en": "Qty", "fr": "Qte", "de": "Menge"},
    "增减配置": {"en": "ADDITIONS / DEDUCTIONS", "fr": "AJOUTS / DEDUCTIONS", "de": "ZUSATZ- / ABWAHLKONFIGURATION"},
    "类型": {"en": "Type", "fr": "Type", "de": "Typ"},
    "部件代号": {"en": "Part code", "fr": "Code piece", "de": "Teilecode"},
    "增减部件": {"en": "Part", "fr": "Piece", "de": "Teil"},
    "交易条款及其他信息": {"en": "TRADE CLAUSE & OTHER INFORMATION", "fr": "CONDITIONS COMMERCIALES ET AUTRES INFORMATIONS", "de": "HANDELSBEDINGUNGEN UND WEITERE INFORMATIONEN"},
    "付款方式": {"en": "Payment item", "fr": "Modalites de paiement", "de": "Zahlungsbedingungen"},
    "交货期": {"en": "Delivery time", "fr": "Delai de livraison", "de": "Lieferzeit"},
    "报价有效期": {"en": "Quotation validity", "fr": "Validite de l'offre", "de": "Angebotsgueltigkeit"},
    "运输方案": {"en": "Transportation", "fr": "Transport", "de": "Transport"},
    "质保期": {"en": "Warranty", "fr": "Garantie", "de": "Garantie"},
    "其他": {"en": "Others", "fr": "Autres", "de": "Sonstiges"},
    "未导入该机型配置表。": {"en": "No configuration workbook imported for this model.", "fr": "Aucun classeur de configuration importe pour ce modele.", "de": "Keine Konfigurationsarbeitsmappe fuer dieses Modell importiert."},
    "未选择增减配置。": {"en": "No additions or deductions selected.", "fr": "Aucun ajout ou deduction selectionne.", "de": "Keine Zusatz- oder Abwahlkonfiguration ausgewaehlt."},
    "增配": {"en": "Addition", "fr": "Ajout", "de": "Zusatz"},
    "减配": {"en": "Deduction", "fr": "Deduction", "de": "Abwahl"},
    "平臂塔机": {"en": "Hammerhead tower crane", "fr": "Grue a tour a fleche horizontale", "de": "Obendreher mit horizontalem Ausleger"},
    "动臂塔机": {"en": "Luffing tower crane", "fr": "Grue a tour a fleche relevable", "de": "Wippausleger-Turmdrehkran"},
    "风电塔机": {"en": "Wind power tower crane", "fr": "Grue a tour pour eoliennes", "de": "Windkraft-Turmdrehkran"},
    "平头动臂塔机": {"en": "Flat-top luffing tower crane", "fr": "Grue a fleche relevable a tete plate", "de": "Flat-Top-Wippauslegerkran"},
    "桅杆吊塔机": {"en": "Derrick tower crane", "fr": "Grue derrick", "de": "Derrickkran"},
    "塔式起重机": {"en": "Tower crane", "fr": "Grue a tour", "de": "Turmdrehkran"},
    "中塔": {"en": "Medium tower crane", "fr": "Grue moyenne", "de": "Mittlerer Turmdrehkran"},
    "小塔": {"en": "Small tower crane", "fr": "Petite grue", "de": "Kleiner Turmdrehkran"},
    "圆榫节": {"en": "Round-tenon mast section", "fr": "Element de mat a tenon rond", "de": "Rundzapfen-Mastsegment"},
    "螺栓节": {"en": "Bolted mast section", "fr": "Element de mat boulonne", "de": "Verschraubtes Mastsegment"},
    "支腿固定式": {"en": "Fixing-angle stationary", "fr": "Fixe sur pieds d'ancrage", "de": "Stationaer mit Fundamentankern"},
    "底架固定式": {"en": "Base-frame stationary", "fr": "Fixe sur chassis de base", "de": "Stationaer mit Grundrahmen"},
    "基础": {"en": "Foundation", "fr": "Fondation", "de": "Fundament"},
    "选配部件": {"en": "Optional parts", "fr": "Pieces optionnelles", "de": "Optionale Teile"},
    "平衡重": {"en": "Counterweight", "fr": "Contrepoids", "de": "Gegengewicht"},
    "空调": {"en": "Air conditioner", "fr": "Climatiseur", "de": "Klimaanlage"},
    "起重臂": {"en": "Jib", "fr": "Fleche", "de": "Ausleger"},
    "平衡臂": {"en": "Counter-jib", "fr": "Contre-fleche", "de": "Gegenausleger"},
    "司机室": {"en": "Cabin", "fr": "Cabine", "de": "Kabine"},
    "上支座": {"en": "Upper support", "fr": "Support superieur", "de": "Oberer Traeger"},
    "回转支承": {"en": "Slewing bearing", "fr": "Couronne d'orientation", "de": "Drehverbindung"},
    "下支座": {"en": "Lower support", "fr": "Support inferieur", "de": "Unterer Traeger"},
    "吊钩": {"en": "Hook", "fr": "Crochet", "de": "Haken"},
    "吊钩组": {"en": "Hook block", "fr": "Moufle a crochet", "de": "Hakenflasche"},
    "标准节": {"en": "Mast section", "fr": "Element de mat", "de": "Mastsegment"},
    "圆榫标准节": {"en": "Round-tenon mast section", "fr": "Element de mat a tenon rond", "de": "Rundzapfen-Mastsegment"},
    "底架": {"en": "Base frame", "fr": "Chassis de base", "de": "Grundrahmen"},
    "支腿组件": {"en": "Support leg assembly", "fr": "Ensemble de pieds de support", "de": "Stuetzbeinbaugruppe"},
    "基础预埋": {"en": "Embedded foundation parts", "fr": "Pieces encastrees de fondation", "de": "Einbauteile fuer Fundament"},
    "机构": {"en": "Mechanism", "fr": "Mecanisme", "de": "Antrieb"},
    "钢丝绳": {"en": "Wire rope", "fr": "Cable acier", "de": "Stahldrahtseil"},
    "电控总成": {"en": "Electric control assembly", "fr": "Ensemble de commande electrique", "de": "Elektrische Steuerungseinheit"},
    "安全钢丝绳": {"en": "Safety wire rope", "fr": "Cable de securite", "de": "Sicherheitsdrahtseil"},
    "起升机构": {"en": "Hoisting mechanism", "fr": "Mecanisme de levage", "de": "Hubwerk"},
    "回转机构": {"en": "Slewing mechanism", "fr": "Mecanisme d'orientation", "de": "Drehwerk"},
    "变幅机构": {"en": "Trolleying mechanism", "fr": "Mecanisme de variation de portee", "de": "Katzfahrwerk"},
    "起升钢丝绳": {"en": "Hoisting wire rope", "fr": "Cable de levage", "de": "Hubseil"},
    "变幅钢丝绳": {"en": "Trolleying wire rope", "fr": "Cable de variation de portee", "de": "Katzfahrseil"},
    "起升机构带插头电缆包": {"en": "Hoisting mechanism plug cable package", "fr": "Kit cable avec fiche du mecanisme de levage", "de": "Steckerkabelsatz fuer Hubwerk"},
    "变幅机构带插头电缆包": {"en": "Trolleying mechanism plug cable package", "fr": "Kit cable avec fiche du mecanisme de variation de portee", "de": "Steckerkabelsatz fuer Katzfahrwerk"},
    "回转机构带插头电缆包": {"en": "Slewing mechanism plug cable package", "fr": "Kit cable avec fiche du mecanisme d'orientation", "de": "Steckerkabelsatz fuer Drehwerk"},
    "4.5m城市底架": {"en": "4.5m city base frame", "fr": "Chassis urbain 4.5 m", "de": "4,5-m-Stadtgrundrahmen"},
    "平衡重(24.6T)": {"en": "Counterweight (24.6t)", "fr": "Contrepoids (24,6 t)", "de": "Gegengewicht (24,6 t)"},
    "平衡重(26.4T)": {"en": "Counterweight (26.4t)", "fr": "Contrepoids (26,4 t)", "de": "Gegengewicht (26,4 t)"},
    "矮标准节": {"en": "Short mast section", "fr": "Element de mat court", "de": "Kurzes Mastsegment"},
    "分体式空调组件": {"en": "Split air conditioner assembly", "fr": "Ensemble climatiseur split", "de": "Split-Klimageraet-Baugruppe"},
    "右交": {"en": "right lay", "fr": "toronnage droit", "de": "rechtsgaengig"},
    "左交": {"en": "left lay", "fr": "toronnage gauche", "de": "linksgaengig"},
    "合同签订后支付30%作为定金，发货前付清剩余70%合同款。": {
        "en": "30% by T/T as prepayment after contract signing, and the remaining 70% before delivery.",
        "fr": "30 % par virement T/T a titre d'acompte apres la signature du contrat, et les 70 % restants avant la livraison.",
        "de": "30 % per T/T als Anzahlung nach Vertragsunterzeichnung und die restlichen 70 % vor Lieferung.",
    },
    "收到定金后90日内发货。": {
        "en": "Delivery within 90 days after receipt of prepayment.",
        "fr": "Livraison dans les 90 jours apres reception de l'acompte.",
        "de": "Lieferung innerhalb von 90 Tagen nach Eingang der Anzahlung.",
    },
    "30天": {"en": "30 days", "fr": "30 jours", "de": "30 Tage"},
    "以最终发运方案为准。": {
        "en": "Subject to the final shipment plan.",
        "fr": "Sous reserve du plan d'expedition final.",
        "de": "Massgeblich ist der endgueltige Versandplan.",
    },
    "自提单之日起保修期：钢结构12个月；机构12个月；电气部件12个月。易损件除外。": {
        "en": "Warranty period from the date of B/L: 12 months for steel structure, 12 months for mechanisms, and 12 months for electric parts. Wearing parts are excluded.",
        "fr": "Periode de garantie a compter de la date du connaissement: 12 mois pour la structure acier, 12 mois pour les mecanismes et 12 mois pour les composants electriques. Les pieces d'usure sont exclues.",
        "de": "Garantiezeit ab Datum des Konnossements: 12 Monate fuer Stahlkonstruktion, 12 Monate fuer Mechanismen und 12 Monate fuer elektrische Teile. Verschleissteile sind ausgeschlossen.",
    },
    "报价含首次安装指导服务费用。": {
        "en": "The quotation includes the first installation guidance service fee.",
        "fr": "Le devis comprend les frais du premier service de guidage d'installation.",
        "de": "Das Angebot enthaelt die Kosten fuer die erstmalige Montageanleitung.",
    },
    "上海港": {"en": "Shanghai Port", "fr": "Port de Shanghai", "de": "Hafen Shanghai"},
    "中联建起（ZOOMLION Construction Hoisting Machinery Company）": {
        "en": "ZOOMLION Construction Hoisting Machinery Company",
        "fr": "ZOOMLION Construction Hoisting Machinery Company",
        "de": "ZOOMLION Construction Hoisting Machinery Company",
    },
    "中联建起": {
        "en": "ZOOMLION Construction Hoisting Machinery Company",
        "fr": "ZOOMLION Construction Hoisting Machinery Company",
        "de": "ZOOMLION Construction Hoisting Machinery Company",
    },
    "中联重科建筑起重机械公司": {
        "en": "ZOOMLION Construction Hoisting Machinery Company",
        "fr": "ZOOMLION Construction Hoisting Machinery Company",
        "de": "ZOOMLION Construction Hoisting Machinery Company",
    },
    "中联重科建筑起重机械有限公司": {
        "en": "ZOOMLION Construction Hoisting Machinery Co., Ltd.",
        "fr": "ZOOMLION Construction Hoisting Machinery Co., Ltd.",
        "de": "ZOOMLION Construction Hoisting Machinery Co., Ltd.",
    },
    "中联塔机配置确认及报价单生成软件 V1.0": {
        "en": "ZOOMLION Tower Crane Configuration Confirmation and Quotation Generator V1.0",
        "fr": "Logiciel de liste de configuration et devis pour grues a tour ZOOMLION",
        "de": "ZOOMLION Software fuer Turmdrehkran-Konfigurationslisten und Angebote",
    },
    "配置清单": {"en": "Configuration list", "fr": "Liste de configuration", "de": "Konfigurationsliste"},
    "基本配置清单": {"en": "Basic configuration list", "fr": "Liste de configuration de base", "de": "Grundkonfigurationsliste"},
    "增减配清单": {"en": "Addition / deduction list", "fr": "Liste des ajouts / deductions", "de": "Zusatz- / Abwahlliste"},
    "配置及增减配清单": {"en": "Configuration and addition / deduction list", "fr": "Liste de configuration et ajouts / deductions", "de": "Konfiguration und Zusatz- / Abwahlliste"},
    "配置表": {"en": "configuration list", "fr": "liste de configuration", "de": "Konfigurationsliste"},
    "平头塔式起重机配置表": {"en": "flat-top tower crane configuration list", "fr": "liste de configuration de grue a tour a tete plate", "de": "Konfigurationsliste fuer Flat-Top-Turmdrehkran"},
    "平头塔式起重机": {"en": "flat-top tower crane", "fr": "grue a tour a tete plate", "de": "Flat-Top-Turmdrehkran"},
    "版本号": {"en": "Version", "fr": "Version", "de": "Version"},
    "发布日期": {"en": "Release date", "fr": "Date de publication", "de": "Veroeffentlichungsdatum"},
    "序号": {"en": "No.", "fr": "No.", "de": "Nr."},
    "组成": {"en": "Module", "fr": "Module", "de": "Modul"},
    "产品模块": {"en": "Product module", "fr": "Module produit", "de": "Produktmodul"},
    "名称": {"en": "Name", "fr": "Nom", "de": "Name"},
    "编码": {"en": "Item No.", "fr": "Reference", "de": "Artikelnummer"},
    "编 码": {"en": "Item No.", "fr": "Reference", "de": "Artikelnummer"},
    "标配": {"en": "Standard", "fr": "Standard", "de": "Standard"},
    "选配": {"en": "Optional", "fr": "Optionnel", "de": "Optional"},
    "不配": {"en": "Not included", "fr": "Non inclus", "de": "Nicht enthalten"},
    "基价": {"en": "Base price", "fr": "Prix de base", "de": "Grundpreis"},
    "不含税": {"en": "tax excluded", "fr": "hors taxe", "de": "ohne Steuer"},
    "FOB基价（不含税）": {"en": "FOB base price (tax excluded)", "fr": "Prix de base FOB (hors taxe)", "de": "FOB-Grundpreis (ohne Steuer)"},
    "“●”标配 “○”选配 “-”不配": {
        "en": "\"●\" standard  \"○\" optional  \"-\" not included",
        "fr": "\"●\" standard  \"○\" optionnel  \"-\" non inclus",
        "de": "\"●\" Standard  \"○\" Optional  \"-\" nicht enthalten",
    },
    "产品型号：": {"en": "Model: ", "fr": "Modele: ", "de": "Modell: "},
    "安装形式：": {"en": "Installation form: ", "fr": "Type d'installation: ", "de": "Aufstellungsart: "},
    "来源配置表：": {"en": "Source workbook: ", "fr": "Classeur source: ", "de": "Quellarbeitsmappe: "},
    "日期": {"en": "Date", "fr": "Date", "de": "Datum"},
    "LTC选配指导文件": {"en": "LTC Optional Parts Guide", "fr": "Guide des options LTC", "de": "LTC-Leitfaden fuer optionale Teile"},
    "部件": {"en": "Component", "fr": "Composant", "de": "Komponente"},
    "项目": {"en": "Item", "fr": "Article", "de": "Position"},
    "支腿 固定式": {"en": "Fixing-angle stationary", "fr": "Fixe sur pieds d'ancrage", "de": "Stationaer mit Fundamentankern"},
    "底架 固定式": {"en": "Base-frame stationary", "fr": "Fixe sur chassis de base", "de": "Stationaer mit Grundrahmen"},
    "支腿": {"en": "Fixing angle", "fr": "Pied d'ancrage", "de": "Fundamentanker"},
    "螺栓固定式": {"en": "Bolted stationary", "fr": "Fixe boulonne", "de": "Stationaer verschraubt"},
    "预埋支腿式": {"en": "Embedded fixing-angle type", "fr": "Type pieds d'ancrage encastres", "de": "Typ mit einbetonierten Fundamentankern"},
    "固定式": {"en": "Stationary", "fr": "Fixe", "de": "Stationaer"},
    "附着式": {"en": "Anchored", "fr": "Ancre", "de": "Verankert"},
    "内爬式": {"en": "Internal climbing", "fr": "Grimpant interieur", "de": "Innenkletternd"},
    "行走式": {"en": "Travelling", "fr": "Mobile sur rails", "de": "Fahrbar"},
    "上装总成": {"en": "Upper structure assembly", "fr": "Ensemble superstructure", "de": "Oberwagenbaugruppe"},
    "塔身": {"en": "Mast", "fr": "Mat", "de": "Mast"},
    "总成": {"en": "Assembly", "fr": "Ensemble", "de": "Baugruppe"},
    "塔身 总成": {"en": "Mast assembly", "fr": "Ensemble de mat", "de": "Mastbaugruppe"},
    "塔身总成": {"en": "Mast assembly", "fr": "Ensemble de mat", "de": "Mastbaugruppe"},
    "安全": {"en": "Safety", "fr": "Securite", "de": "Sicherheit"},
    "装置": {"en": "Device", "fr": "Dispositif", "de": "Vorrichtung"},
    "安全 装置": {"en": "Safety device", "fr": "Dispositif de securite", "de": "Sicherheitsvorrichtung"},
    "安全装置": {"en": "Safety device", "fr": "Dispositif de securite", "de": "Sicherheitsvorrichtung"},
    "标牌": {"en": "Nameplates", "fr": "Plaques signaletiques", "de": "Schilder"},
    "机构及钢丝绳": {"en": "Mechanisms and wire ropes", "fr": "Mecanismes et cables acier", "de": "Mechanismen und Stahldrahtseile"},
    "电控系统": {"en": "Electric control system", "fr": "Systeme de commande electrique", "de": "Elektrisches Steuerungssystem"},
    "电气部件": {"en": "Electric parts", "fr": "Composants electriques", "de": "Elektrische Teile"},
    "驱动机构": {"en": "Drive mechanism", "fr": "Mecanisme d'entrainement", "de": "Antriebsmechanismus"},
    "上装附件": {"en": "Upper structure accessories", "fr": "Accessoires de superstructure", "de": "Oberwagenzubehoer"},
    "绳夹": {"en": "Wire rope clip", "fr": "Serre-cable", "de": "Seilklemme"},
    "信号接收器安装座": {"en": "Signal receiver mounting bracket", "fr": "Support de montage du recepteur de signal", "de": "Montagehalter fuer Signalempfaenger"},
    "回转制动器显示灯安装座": {"en": "Slewing brake indicator mounting bracket", "fr": "Support du voyant de frein d'orientation", "de": "Halter fuer Drehbrems-Kontrollleuchte"},
    "三色灯安装座": {"en": "Three-color light mounting bracket", "fr": "Support du feu tricolore", "de": "Halter fuer dreifarbige Leuchte"},
    "载重小车": {"en": "Trolley", "fr": "Chariot", "de": "Laufkatze"},
    "圆榫基础节": {"en": "Round-tenon base mast section", "fr": "Element de base a tenon rond", "de": "Rundzapfen-Grundmastsegment"},
    "RA外挂式平台": {"en": "RA external platform", "fr": "Plateforme externe RA", "de": "RA-Aussenplattform"},
    "RA支腿座": {"en": "RA fixing-angle seat", "fr": "Support de pied d'ancrage RA", "de": "RA-Fundamentankeraufnahme"},
    "支腿座": {"en": "Fixing-angle seat", "fr": "Support de pied d'ancrage", "de": "Fundamentankeraufnahme"},
    "开关安装盒": {"en": "Switch mounting box", "fr": "Boitier de montage d'interrupteur", "de": "Schaltermontagekasten"},
    "起重量传感器总成(含线缆)": {"en": "Load sensor assembly (with cable)", "fr": "Ensemble capteur de charge (avec cable)", "de": "Lastsensorbaugruppe (mit Kabel)"},
    "含线缆": {"en": "with cable", "fr": "avec cable", "de": "mit Kabel"},
    "起重量限制器拉环": {"en": "Load moment limiter pull ring", "fr": "Anneau de traction du limiteur de charge", "de": "Zugring fuer Lastbegrenzer"},
    "回转限位装置": {"en": "Slewing limit device", "fr": "Dispositif de limitation d'orientation", "de": "Drehbegrenzungsvorrichtung"},
    "起升绝对值编码器": {"en": "Hoisting absolute encoder", "fr": "Codeur absolu de levage", "de": "Absolutwertgeber fuer Hubwerk"},
    "回转绝对值编码器": {"en": "Slewing absolute encoder", "fr": "Codeur absolu d'orientation", "de": "Absolutwertgeber fuer Drehwerk"},
    "变幅绝对值编码器": {"en": "Trolleying absolute encoder", "fr": "Codeur absolu de variation de portee", "de": "Absolutwertgeber fuer Katzfahrwerk"},
    "起重特性牌": {"en": "Load chart plate", "fr": "Plaque des caracteristiques de charge", "de": "Lastdiagrammschild"},
    "CE铭牌": {"en": "CE nameplate", "fr": "Plaque signaletique CE", "de": "CE-Typenschild"},
    "CE铭牌安装架": {"en": "CE nameplate mounting bracket", "fr": "Support de plaque CE", "de": "Halter fuer CE-Typenschild"},
    "幅度指示牌": {"en": "Radius indicator plate", "fr": "Plaque d'indication de portee", "de": "Ausladungsanzeige"},
    "司机安全操作规程牌": {"en": "Operator safety instruction plate", "fr": "Plaque des consignes de securite operateur", "de": "Sicherheitsanweisungsschild fuer Kranfuehrer"},
    "序号牌": {"en": "Serial number plate", "fr": "Plaque de numero de serie", "de": "Seriennummernschild"},
    "司机室电源箱总成": {"en": "Cabin power box assembly", "fr": "Ensemble coffret d'alimentation cabine", "de": "Kabinen-Stromkastenbaugruppe"},
    "司机室插头电缆包": {"en": "Cabin plug cable package", "fr": "Kit cable avec fiche cabine", "de": "Kabinen-Steckerkabelsatz"},
    "无线遥控附件包": {"en": "Wireless remote-control accessory package", "fr": "Kit accessoires de radiocommande", "de": "Zubehoerpaket fuer Funkfernsteuerung"},
    "电源箱总成": {"en": "Power box assembly", "fr": "Ensemble coffret d'alimentation", "de": "Stromkastenbaugruppe"},
    "风速仪": {"en": "Anemometer", "fr": "Anemometre", "de": "Windmesser"},
    "风速仪电缆": {"en": "Anemometer cable", "fr": "Cable d'anemometre", "de": "Windmesserkabel"},
    "电阻器": {"en": "Resistor", "fr": "Resistance", "de": "Widerstand"},
    "主电缆": {"en": "Main cable", "fr": "Cable principal", "de": "Hauptkabel"},
    "4.5m城市底架压重": {"en": "4.5m city base-frame ballast", "fr": "Lest de chassis urbain 4,5 m", "de": "Ballast fuer 4,5-m-Stadtgrundrahmen"},
    "DJ-4.5m压重": {"en": "DJ-4.5m ballast", "fr": "Lest DJ-4,5 m", "de": "DJ-4,5-m-Ballast"},
    "平衡重框": {"en": "Counterweight frame", "fr": "Cadre de contrepoids", "de": "Gegengewichtsrahmen"},
    "防碰撞系统": {"en": "Anti-collision system", "fr": "Systeme anticollision", "de": "Antikollisionssystem"},
    "1.5P有氟分体式空调": {"en": "1.5P fluorinated split air conditioner", "fr": "Climatiseur split fluore 1,5P", "de": "1,5P fluorierte Split-Klimaanlage"},
    "预埋支腿定位框": {"en": "Embedded fixing-angle positioning frame", "fr": "Cadre de positionnement des pieds d'ancrage encastres", "de": "Positionierrahmen fuer einbetonierte Fundamentanker"},
    "Autec无线遥控器": {"en": "Autec wireless remote control", "fr": "Radiocommande Autec", "de": "Autec-Funkfernsteuerung"},
    "爬升 部件包": {"en": "Climbing package", "fr": "Kit de grimpement", "de": "Kletterpaket"},
    "爬升": {"en": "Climbing", "fr": "Grimpement", "de": "Klettern"},
    "部件包": {"en": "Package", "fr": "Kit", "de": "Paket"},
    "爬升包": {"en": "Climbing package", "fr": "Kit de grimpement", "de": "Kletterpaket"},
    "爬升架": {"en": "Climbing frame", "fr": "Cage de grimpement", "de": "Kletterrahmen"},
    "顶升机构": {"en": "Jacking mechanism", "fr": "Mecanisme de hissage", "de": "Hydraulisches Kletterwerk"},
    "顶升油缸": {"en": "Jacking cylinder", "fr": "Verin de hissage", "de": "Kletterzylinder"},
    "泵站": {"en": "Pump station", "fr": "Station de pompage", "de": "Pumpenstation"},
    "胶管总成": {"en": "Hose assembly", "fr": "Ensemble flexible", "de": "Schlauchbaugruppe"},
    "过渡节": {"en": "Transition mast section", "fr": "Element de transition", "de": "Uebergangsmastsegment"},
    "引进系统": {"en": "Mast introduction system", "fr": "Systeme d'introduction des elements", "de": "Einfuehrsystem fuer Mastsegmente"},
    "安装平台": {"en": "Installation platform", "fr": "Plateforme d'installation", "de": "Montageplattform"},
    "中央集电环配套物资": {"en": "Central collector-ring supporting parts", "fr": "Pieces associees au collecteur central", "de": "Zubehoer fuer zentralen Schleifring"},
    "中央集电环包": {"en": "Central collector-ring package", "fr": "Kit collecteur central", "de": "Zentrales Schleifringpaket"},
    "中央集电环": {"en": "Central collector ring", "fr": "Collecteur central", "de": "Zentraler Schleifring"},
    "W45回转绝对值编码器电缆": {"en": "W45 slewing absolute encoder cable", "fr": "Cable du codeur absolu d'orientation W45", "de": "W45-Kabel fuer Drehabsolutwertgeber"},
    "圆榫非标节": {"en": "Round-tenon non-standard mast section", "fr": "Element de mat non standard a tenon rond", "de": "Nicht standardisiertes Rundzapfen-Mastsegment"},
    "RA-3半节": {"en": "RA-3 half mast section", "fr": "Demi-element RA-3", "de": "RA-3 Halbsegment"},
    "RA-2半节": {"en": "RA-2 half mast section", "fr": "Demi-element RA-2", "de": "RA-2 Halbsegment"},
    "电动扒杆包": {"en": "Electric derrick package", "fr": "Kit derrick electrique", "de": "Elektrisches Derrickpaket"},
    "1T电动扒杆": {"en": "1t electric derrick", "fr": "Derrick electrique 1 t", "de": "1-t-Elektroderrick"},
    "电动扒杆电控系统": {"en": "Electric derrick control system", "fr": "Systeme de commande du derrick electrique", "de": "Steuerungssystem fuer Elektroderrick"},
}


def ensure_dirs():
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


def default_database():
    return {
        "products": {},
        "forms_by_code": {},
        "components_by_model": {},
        "config_exports_by_model": {},
        "deleted_config_model_keys": [],
        "published_config_model_keys": [],
        "rejected_config_model_keys": [],
        "options_by_code": {},
        "change_log": [],
        "config_data_reset_version": CONFIG_DATA_RESET_VERSION,
    }


def apply_database_migrations(db, reset_config_data=False):
    changed = False
    if reset_config_data:
        db["components_by_model"] = {}
        db["config_exports_by_model"] = {}
        db["deleted_config_model_keys"] = []
        db["published_config_model_keys"] = []
        db["rejected_config_model_keys"] = []
        db["config_data_reset_version"] = CONFIG_DATA_RESET_VERSION
        changed = True
    return changed


def load_database():
    ensure_dirs()
    source = DB_FILE if DB_FILE.exists() else INITIAL_DB_FILE
    if not source.exists():
        return default_database()
    with source.open("r", encoding="utf-8") as f:
        db = json.load(f)
    reset_config_data = int(db.get("config_data_reset_version", 0) or 0) < CONFIG_DATA_RESET_VERSION
    base = default_database()
    base.update(db)
    changed = apply_database_migrations(base, reset_config_data)
    if changed and source == DB_FILE:
        with DB_FILE.open("w", encoding="utf-8") as f:
            json.dump(base, f, ensure_ascii=False, indent=2)
    return base


def save_database(db):
    ensure_dirs()
    with DB_FILE.open("w", encoding="utf-8") as f:
        json.dump(db, f, ensure_ascii=False, indent=2)


def append_change_log(operator_id, operation_time, content):
    db = load_database()
    db.setdefault("change_log", []).append(
        {
            "time": clean_text(operation_time) or datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "operator_id": clean_text(operator_id),
            "content": clean_text(content),
        }
    )
    save_database(db)


def default_user_settings():
    return {
        "header_address": "Zoomlion Smart City Headquarters Building, No.613 NaqiuRoad, WangCheng District, Changsha, Hunan, China",
        "customer_name": "",
        "quote_company": "中联重科建筑起重机械有限公司",
        "quote_person": "Lewis",
        "quote_phone": "+86 123456789",
        "quote_email": "liuyuchi@zoomlion.com",
    }


def load_user_settings():
    ensure_dirs()
    settings = default_user_settings()
    if SETTINGS_FILE.exists():
        try:
            with SETTINGS_FILE.open("r", encoding="utf-8") as f:
                saved = json.load(f)
            if isinstance(saved, dict):
                settings.update({key: clean_text(value) for key, value in saved.items() if key in settings})
        except Exception:
            pass
    if settings.get("quote_company") == "中联重科建筑起重机械公司":
        settings["quote_company"] = "中联重科建筑起重机械有限公司"
    if settings.get("quote_email") == "Lewisliu@zoomlion.com":
        settings["quote_email"] = "liuyuchi@zoomlion.com"
    if not settings.get("quote_email"):
        settings["quote_email"] = "liuyuchi@zoomlion.com"
    return settings


def save_user_settings(settings):
    ensure_dirs()
    with SETTINGS_FILE.open("w", encoding="utf-8") as f:
        json.dump(settings, f, ensure_ascii=False, indent=2)


def clean_text(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def clean_config_text(value):
    return clean_text(value).replace("图", "")


def code_or_slash(value):
    return clean_text(value) or "/"


def join_display_parts(parts):
    return "、".join(clean_text(part) for part in parts if clean_text(part))


def compact_display_text(value):
    return re.sub(r"\s+", "", clean_text(value))


def normalize_key(value):
    return re.sub(r"\s+", "", clean_text(value)).replace("（", "(").replace("）", ")").upper()


def normalize_form(value):
    return re.sub(r"\s+", "", clean_text(value))


def is_visible_product_model(model_name):
    model = normalize_key(model_name)
    return not (model.startswith("D") or model.startswith("ZT"))


def format_number(value):
    text = clean_text(value).replace(",", "")
    try:
        number = float(text)
    except ValueError:
        return clean_text(value)
    if number.is_integer():
        return f"{int(number):,}"
    return f"{number:,.2f}".rstrip("0").rstrip(".")


def parse_price_value(value):
    text = clean_text(value).replace(",", "")
    if not text:
        return 0.0
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return 0.0
    try:
        return float(match.group())
    except ValueError:
        return 0.0


def format_price_value(value):
    sign = "-" if value < 0 else ""
    return sign + format_number(abs(value))


def parse_config_description(description):
    result = {}
    for part in clean_text(description).split(","):
        if ":" in part:
            key, value = part.split(":", 1)
            result[key.strip()] = value.strip()
    return {
        "tower_size": result.get("类型", ""),
        "jib_length": result.get("臂长(M)", ""),
        "tonnage": result.get("吨数(T)", ""),
        "mast_type": result.get("塔身种类", ""),
    }


def detect_tower_type(model_name):
    model = normalize_key(model_name)
    if model.startswith("LW"):
        return {"zh": "风电塔机", "en": "Wind power tower crane"}
    if model.startswith("RL"):
        return {"zh": "平头动臂塔机", "en": "Flat-top luffing tower crane"}
    if model.startswith("LR"):
        return {"zh": "桅杆吊塔机", "en": "Derrick tower crane"}
    if model.startswith("L"):
        return {"zh": "动臂塔机", "en": "Luffing tower crane"}
    if model.startswith(("WA", "W", "T", "R")):
        return {"zh": "平臂塔机", "en": "Hammerhead tower crane"}
    return {"zh": "塔式起重机", "en": "Tower crane"}


def detect_max_load(model_name):
    match = re.search(r"-(\d+(?:\.\d+)?)", clean_text(model_name))
    if not match:
        return ""
    value = match.group(1)
    return f"{value}t"


def import_price_file(path):
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    db = load_database()
    imported_products = set()
    option_count = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        product_group = clean_text(row[3] if len(row) > 3 else "")
        if product_group != "B1":
            continue
        model = clean_text(row[4] if len(row) > 4 else "")
        config_desc = clean_text(row[5] if len(row) > 5 else "")
        product_code = clean_text(row[6] if len(row) > 6 else "")
        if not model or not product_code:
            continue

        parsed = parse_config_description(config_desc)
        db["products"][model] = {
            "model": model,
            "model_key": normalize_key(model),
            "product_code": product_code,
            "config_description": config_desc,
            "tower_size": parsed["tower_size"],
            "default_jib_length": parsed["jib_length"],
            "rated_tonnage": parsed["tonnage"],
            "mast_type": parsed["mast_type"],
            "tower_type": detect_tower_type(model),
            "max_load": detect_max_load(model),
        }
        imported_products.add(model)

        option = {
            "category_code": clean_text(row[8] if len(row) > 8 else ""),
            "category": clean_text(row[9] if len(row) > 9 else ""),
            "part_code": clean_text(row[10] if len(row) > 10 else ""),
            "part_name": clean_text(row[11] if len(row) > 11 else ""),
            "material_code": clean_text(row[12] if len(row) > 12 else ""),
            "material_description": clean_text(row[13] if len(row) > 13 else ""),
            "change_type_code": clean_text(row[14] if len(row) > 14 else ""),
            "add_price": format_number(row[15] if len(row) > 15 else ""),
            "deduct_price": format_number(row[16] if len(row) > 16 else ""),
        }
        if option["part_name"] or option["part_code"]:
            db["options_by_code"].setdefault(product_code, [])
            dedupe = "|".join([option["category_code"], option["part_code"], option["part_name"], option["change_type_code"]])
            if not any(item.get("_dedupe") == dedupe for item in db["options_by_code"][product_code]):
                option["_dedupe"] = dedupe
                db["options_by_code"][product_code].append(option)
                option_count += 1

    save_database(db)
    return len(imported_products), option_count


def import_product_cfg_file(path):
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    db = load_database()
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        product_code = clean_text(row[0] if len(row) > 0 else "")
        if not product_code:
            continue
        form = {
            "install_code": clean_text(row[1] if len(row) > 1 else ""),
            "install_form": clean_text(row[2] if len(row) > 2 else ""),
            "height": clean_text(row[9] if len(row) > 9 else ""),
            "jib_length": clean_text(row[10] if len(row) > 10 else ""),
            "rope_capacity": clean_text(row[11] if len(row) > 11 else ""),
            "wire_rope_spec": clean_text(row[12] if len(row) > 12 else ""),
            "fall": clean_text(row[13] if len(row) > 13 else ""),
            "hoisting_height": clean_text(row[14] if len(row) > 14 else ""),
        }
        if not form["install_form"]:
            continue
        forms = db["forms_by_code"].setdefault(product_code, [])
        key = normalize_form(form["install_form"])
        forms[:] = [item for item in forms if normalize_form(item.get("install_form")) != key]
        forms.append(form)
        count += 1
    save_database(db)
    return count


def list_section_bounds(ws, list_type):
    section_rows = []
    for row_idx in range(1, ws.max_row + 1):
        row_text = " ".join(clean_text(ws.cell(row_idx, col).value) for col in range(1, ws.max_column + 1))
        compact = re.sub(r"\s+", "", row_text)
        if not compact:
            continue
        if "增减配清单" in compact or ("增减" in compact and "清单" in compact):
            section_rows.append(("option", row_idx))
        elif "标准配置清单" in compact or ("配置清单" in compact and "增减" not in compact):
            section_rows.append(("basic", row_idx))
    if not section_rows:
        return 1, ws.max_row

    wanted = "option" if list_type == "option" else "basic"
    matching = [item for item in section_rows if item[0] == wanted]
    if not matching:
        return 1, ws.max_row
    start_row = matching[0][1]
    following = [row for _kind, row in section_rows if row > start_row]
    end_row = min(following) - 1 if following else ws.max_row
    return start_row, end_row


def find_config_form_columns(ws, start_row=1, end_row=None):
    end_row = end_row or ws.max_row
    form_row = None
    scan_end = min(end_row, start_row + 14)
    for row_idx in range(start_row, scan_end + 1):
        values = [clean_text(ws.cell(row_idx, col).value) for col in range(1, ws.max_column + 1)]
        row_text = " ".join(values)
        if "标配" in row_text and "选配" in row_text and "不配" in row_text:
            form_row = row_idx + 1
            break
    form_tokens = ["固定式", "附着", "内爬", "行走", "压重"]
    for row_idx in range(start_row, scan_end + 1):
        if form_row:
            break
        values = [clean_text(ws.cell(row_idx, col).value) for col in range(1, ws.max_column + 1)]
        form_columns = [
            col
            for col, value in enumerate(values, start=1)
            if any(token in normalize_form(value) for token in form_tokens)
        ]
        if len(form_columns) >= 2:
            form_row = row_idx
            break
        if len(form_columns) == 1 and form_columns[0] > 1:
            form_row = row_idx
    if not form_row:
        return start_row + 3, {}

    columns = {}
    for col in range(1, ws.max_column + 1):
        value = clean_text(ws.cell(form_row, col).value)
        if not value:
            continue
        if "备注" in value or "更改" in value or "标配" in value or "选配" in value:
            continue
        if any(token in normalize_form(value) for token in form_tokens):
            columns[normalize_form(value)] = col
    return form_row, columns


def classify_component(row_text):
    for item in COMPONENT_KEYWORDS:
        keyword, en = item[0], item[1]
        display_zh = item[2] if len(item) > 2 else keyword
        if keyword in row_text:
            return {"zh": display_zh, "en": en}
    return None


def should_skip_component(row_text):
    text = clean_text(row_text)
    return "插头电缆包" in text or "图" in text


def quantity_from_mark(mark):
    text = clean_text(mark)
    match = re.search(r"\u25cf\s*[\u00d7xX]\s*(\d+(?:\.\d+)?)", text)
    if match:
        return match.group(1)
    return "1"


EXPORT_OPTION_KEYWORDS = ["标准节", "基节", "基础节", "底架", "支腿座", "基础预埋", "支腿组件"]
EXPORT_DROP_KEYWORDS = ["序号", "编码", "备注", "重量", "更改", "价格", "基价"]
EXPORT_SCHEMA_VERSION = 3


def column_label(col):
    label = ""
    while col:
        col, remainder = divmod(col - 1, 26)
        label = chr(65 + remainder) + label
    return label


def available_output_path(path):
    path = Path(path)
    if not path.exists():
        return path
    stamp = datetime.now().strftime("%H%M%S")
    candidate = path.with_name(f"{path.stem}_{stamp}{path.suffix}")
    counter = 2
    while candidate.exists():
        candidate = path.with_name(f"{path.stem}_{stamp}_{counter}{path.suffix}")
        counter += 1
    return candidate


def save_workbook_without_overwrite_conflict(wb, output_path):
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    target = available_output_path(output_path)
    try:
        wb.save(target)
        return target
    except PermissionError:
        fallback = available_output_path(output_path)
        wb.save(fallback)
        return fallback


def categorized_output_path(category, filename):
    folder = OUTPUT_DIR / category
    folder.mkdir(parents=True, exist_ok=True)
    return folder / filename


def available_output_folder(path):
    path = Path(path)
    if not path.exists():
        path.mkdir(parents=True, exist_ok=True)
        return path
    counter = 2
    while True:
        candidate = path.with_name(f"{path.name}_{counter}")
        if not candidate.exists():
            candidate.mkdir(parents=True, exist_ok=True)
            return candidate
        counter += 1


def open_folder(path):
    folder = Path(path).resolve()
    if folder.exists():
        os.startfile(folder)


def open_containing_folder(path):
    folder = Path(path).resolve().parent
    if folder.exists():
        os.startfile(folder)


def column_header_text(ws, col, form_row, start_row=1):
    values = [clean_text(ws.cell(row, col).value) for row in range(start_row, form_row + 1)]
    return " ".join(value for value in values if value)


def export_column_indexes(ws, form_row, form_col, all_form_cols, start_row=1, end_row=None):
    end_row = end_row or ws.max_row
    columns = []
    for col in range(1, ws.max_column + 1):
        header_text = column_header_text(ws, col, form_row, start_row)
        normalized_header = re.sub(r"\s+", "", header_text)
        if any(keyword in normalized_header for keyword in EXPORT_DROP_KEYWORDS):
            continue
        if col in all_form_cols and col != form_col:
            continue
        if not header_text and not any(clean_text(ws.cell(row_idx, col).value) for row_idx in range(form_row + 1, end_row + 1)):
            continue
        columns.append(col)
    return columns


def row_values(ws, row_idx, columns):
    return [clean_text(ws.cell(row_idx, col).value) for col in columns]


def cell_value_with_merge(ws, row_idx, col_idx):
    value = ws.cell(row_idx, col_idx).value
    if value is not None:
        return clean_text(value)
    for merged in ws.merged_cells.ranges:
        if merged.min_row <= row_idx <= merged.max_row and merged.min_col <= col_idx <= merged.max_col:
            return clean_text(ws.cell(merged.min_row, merged.min_col).value)
    return ""


def find_header_column(ws, form_row, keywords, start_row=1):
    for col in range(1, ws.max_column + 1):
        header_text = re.sub(r"\s+", "", column_header_text(ws, col, form_row, start_row))
        if any(keyword in header_text for keyword in keywords):
            return col
    return None


def workbook_bytes_for_model(db, model_name):
    if not is_config_published(db, model_name):
        return None
    for cfg in [find_export_config(db, model_name), find_component_config(db, model_name), find_initial_export_config(model_name)]:
        workbook_base64 = clean_text(cfg.get("workbook_base64", "")) if cfg else ""
        if workbook_base64:
            return base64.b64decode(workbook_base64)
        source_path = find_source_workbook_path(cfg.get("source_file", "")) if cfg else None
        if source_path:
            return source_path.read_bytes()
    return None


def worksheet_for_list_type(wb, list_type):
    if list_type == "option" and len(wb.worksheets) > 1:
        for ws in wb.worksheets[1:]:
            if "增减" in clean_text(ws.title) or "选配" in clean_text(ws.title):
                return ws
        return wb.worksheets[1]
    return wb.worksheets[0]


def matched_form_column(form_columns, install_form):
    form_key = normalize_form(install_form)
    form_col = form_columns.get(form_key)
    if form_col:
        return form_col
    for saved_form, col in form_columns.items():
        if saved_form in form_key or form_key in saved_form:
            return col
    return None


def config_rows_for_form(db, model_name, install_form, list_type="basic"):
    if not is_config_published(db, model_name):
        raise ValueError("当前机型配置表尚未发布，不能读取配置表信息。")
    workbook_bytes = workbook_bytes_for_model(db, model_name)
    if not workbook_bytes:
        db, _cfg = ensure_export_config(db, model_name)
        workbook_bytes = workbook_bytes_for_model(db, model_name)
    if not workbook_bytes:
        raise ValueError("当前机型没有可读取的配置表原始数据。")

    wb = load_workbook(BytesIO(workbook_bytes), data_only=True)
    ws = worksheet_for_list_type(wb, list_type)
    section_start, section_end = list_section_bounds(ws, list_type)
    form_row, form_columns = find_config_form_columns(ws, section_start, section_end)
    form_col = matched_form_column(form_columns, install_form)
    if not form_col:
        raise ValueError("当前安装形式没有匹配到配置表列。")

    comp_col = find_header_column(ws, form_row, ["组成"], section_start)
    part_col = find_header_column(ws, form_row, ["部件"], section_start)
    name_col = find_header_column(ws, form_row, ["名称"], section_start)
    code_col = find_header_column(ws, form_row, ["编码"], section_start)
    model_code_col = find_header_column(ws, form_row, ["代号"], section_start)
    price_col = find_header_column(ws, form_row, ["价格", "基价"], section_start)
    rows = []
    for row_idx in range(form_row + 1, section_end + 1):
        full_row_text = " ".join(clean_text(ws.cell(row_idx, col).value) for col in range(1, ws.max_column + 1))
        if not full_row_text:
            continue
        mark = clean_text(ws.cell(row_idx, form_col).value)
        composition_value = clean_config_text(cell_value_with_merge(ws, row_idx, comp_col)) if comp_col else ""
        component_value = clean_config_text(cell_value_with_merge(ws, row_idx, part_col)) if part_col else composition_value
        rows.append(
            {
                "composition": composition_value,
                "component": component_value,
                "name": clean_config_text(cell_value_with_merge(ws, row_idx, name_col)) if name_col else "",
                "code": clean_config_text(cell_value_with_merge(ws, row_idx, code_col)) if code_col else "",
                "model_code": clean_config_text(cell_value_with_merge(ws, row_idx, model_code_col)) if model_code_col else "",
                "mark": clean_config_text(mark),
                "price": format_number(cell_value_with_merge(ws, row_idx, price_col)) if price_col else "",
                "row_text": clean_config_text(full_row_text),
            }
        )
    return rows


def basic_config_items(db, model_name, install_form):
    return [row for row in config_rows_for_form(db, model_name, install_form) if "●" in row.get("mark", "")]


def option_config_items(db, model_name, install_form):
    rows = config_rows_for_form(db, model_name, install_form, list_type="option")
    items = []
    package_groups = {}
    for row in rows:
        text = " ".join(
            [
                row.get("composition", ""),
                row.get("component", ""),
                row.get("name", ""),
                row.get("code", ""),
                row.get("model_code", ""),
                row.get("row_text", ""),
            ]
        )
        if "电源箱总成" in text:
            continue
        if "○" in row.get("mark", "") or any(keyword in row.get("row_text", "") for keyword in EXPORT_OPTION_KEYWORDS):
            item = dict(row)
            composition = clean_config_text(item.get("composition", ""))
            if compact_display_text(composition) == "选配":
                item["item_display"] = join_display_parts([item.get("name", ""), code_or_slash(item.get("model_code", ""))])
            else:
                item["item_display"] = join_display_parts([item.get("component", ""), item.get("name", ""), code_or_slash(item.get("model_code", ""))])
            if composition and compact_display_text(composition) != "选配":
                package_groups.setdefault(composition, []).append(item)
            else:
                items.append(item)
    for composition, children in package_groups.items():
        price = next((clean_text(child.get("price", "")) for child in children if clean_text(child.get("price", ""))), "")
        display_composition = compact_display_text(composition)
        items.append(
            {
                "composition": composition,
                "component": display_composition,
                "name": display_composition,
                "code": "",
                "model_code": "",
                "mark": "○",
                "price": price,
                "item_display": display_composition,
                "children": children,
            }
        )
    return items


def export_merge_ranges(ws, source_rows, columns):
    row_map = {row_idx: export_idx for export_idx, row_idx in enumerate(source_rows, start=1)}
    col_map = {col_idx: export_idx for export_idx, col_idx in enumerate(columns, start=1)}
    ranges = []
    for merged in ws.merged_cells.ranges:
        rows = list(range(merged.min_row, merged.max_row + 1))
        cols = list(range(merged.min_col, merged.max_col + 1))
        if not all(row in row_map for row in rows):
            continue
        if not all(col in col_map for col in cols):
            continue
        ranges.append(
            {
                "min_row": row_map[merged.min_row],
                "max_row": row_map[merged.max_row],
                "min_col": col_map[merged.min_col],
                "max_col": col_map[merged.max_col],
            }
        )
    return ranges


def build_config_export_rows(ws, form_row, form_columns, start_row=1, end_row=None):
    end_row = end_row or ws.max_row
    all_form_cols = set(form_columns.values())
    export_forms = {}
    for form_key, form_col in form_columns.items():
        columns = export_column_indexes(ws, form_row, form_col, all_form_cols, start_row, end_row)
        form_export_col = columns.index(form_col) if form_col in columns else None
        header_rows = [row_values(ws, row_idx, columns) for row_idx in range(start_row, form_row + 1)]
        basic_rows = []
        basic_source_rows = list(range(start_row, form_row + 1))
        option_rows = []
        option_source_rows = list(range(start_row, form_row + 1))
        for row_idx in range(form_row + 1, end_row + 1):
            full_row_text = " ".join(clean_text(ws.cell(row_idx, col).value) for col in range(1, ws.max_column + 1))
            if not full_row_text:
                continue
            mark = clean_text(ws.cell(row_idx, form_col).value)
            values = row_values(ws, row_idx, columns)
            if "●" in mark:
                basic_rows.append(values)
                basic_source_rows.append(row_idx)
            keyword_match = any(keyword in full_row_text for keyword in EXPORT_OPTION_KEYWORDS)
            if "○" in mark or keyword_match:
                values = list(values)
                if keyword_match and form_export_col is not None:
                    values[form_export_col] = "○"
                option_rows.append(values)
                option_source_rows.append(row_idx)
        export_forms[form_key] = {
            "headers": header_rows,
            "basic_rows": basic_rows,
            "option_rows": option_rows,
            "basic_merges": export_merge_ranges(ws, basic_source_rows, columns),
            "option_merges": export_merge_ranges(ws, option_source_rows, columns),
        }
    return export_forms


def import_tower_config_file(path, model_name=None):
    workbook_bytes = Path(path).read_bytes()
    wb = load_workbook(path, data_only=True)
    ws = worksheet_for_list_type(wb, "basic")
    model_name = clean_text(model_name) or Path(path).stem.replace("（", "(").replace("）", ")")

    basic_start, basic_end = list_section_bounds(ws, "basic")
    form_row, form_columns = find_config_form_columns(ws, basic_start, basic_end)
    components_by_form = {form: [] for form in form_columns}
    export_forms = build_config_export_rows(ws, form_row, form_columns, basic_start, basic_end)

    option_ws = worksheet_for_list_type(wb, "option")
    option_start, option_end = list_section_bounds(option_ws, "option")
    option_form_row, option_form_columns = find_config_form_columns(option_ws, option_start, option_end)
    if option_form_columns:
        option_export_forms = build_config_export_rows(option_ws, option_form_row, option_form_columns, option_start, option_end)
        for form_key, option_data in option_export_forms.items():
            target = export_forms.setdefault(form_key, {})
            target.setdefault("headers", option_data.get("headers", []))
            target.setdefault("basic_rows", [])
            target.setdefault("basic_merges", [])
            target["option_rows"] = option_data.get("option_rows", [])
            target["option_merges"] = option_data.get("option_merges", [])

    name_col = find_header_column(ws, form_row, ["名称"], basic_start)
    model_code_col = find_header_column(ws, form_row, ["代号"], basic_start)

    for row_idx in range(form_row + 1, basic_end + 1):
        source_values = [ws.cell(row_idx, col).value for col in range(1, ws.max_column + 1)]
        row_text = " ".join(clean_text(value) for value in source_values)
        if should_skip_component(row_text):
            continue
        component = classify_component(row_text)
        if not component:
            continue
        part_name = clean_text(ws.cell(row_idx, name_col).value) if name_col else ""
        code = clean_text(ws.cell(row_idx, model_code_col).value) if model_code_col else ""
        part_name = part_name or component["zh"]
        for form_key, col in form_columns.items():
            mark = clean_text(ws.cell(row_idx, col).value)
            if "●" not in mark:
                continue
            components_by_form[form_key].append(
                {
                    "component_zh": component["zh"],
                    "component_en": component["en"],
                    "part_name": part_name,
                    "code": code,
                    "quantity": quantity_from_mark(mark),
                }
            )

    db = load_database()
    model_key = normalize_key(model_name)
    db["deleted_config_model_keys"] = [
        key for key in db.get("deleted_config_model_keys", []) if normalize_key(key) != model_key
    ]
    published_keys = {normalize_key(key) for key in db.get("published_config_model_keys", []) if clean_text(key)}
    published_keys.discard(model_key)
    db["published_config_model_keys"] = sorted(published_keys)
    rejected_keys = {normalize_key(key) for key in db.get("rejected_config_model_keys", []) if clean_text(key)}
    rejected_keys.discard(model_key)
    db["rejected_config_model_keys"] = sorted(rejected_keys)
    old_sources = []
    for key, cfg in list(db["components_by_model"].items()):
        if cfg.get("model_key") == model_key or normalize_key(key) == model_key:
            old_sources.append(clean_text(cfg.get("source_file")) or key)
            del db["components_by_model"][key]
    for key, cfg in list(db["config_exports_by_model"].items()):
        if cfg.get("model_key") == model_key or normalize_key(key) == model_key:
            del db["config_exports_by_model"][key]
    db["components_by_model"][model_name] = {
        "model": model_name,
        "model_key": model_key,
        "source_file": Path(path).name,
        "workbook_base64": base64.b64encode(workbook_bytes).decode("ascii"),
        "forms": components_by_form,
    }
    db["config_exports_by_model"][model_name] = {
        "model": model_name,
        "model_key": model_key,
        "source_file": Path(path).name,
        "workbook_base64": base64.b64encode(workbook_bytes).decode("ascii"),
        "export_schema_version": EXPORT_SCHEMA_VERSION,
        "forms": export_forms,
    }
    save_database(db)
    return model_name, sum(len(items) for items in components_by_form.values()), old_sources


def delete_tower_config(model_name):
    db = load_database()
    model_key = normalize_key(model_name)
    removed_sources = []
    removed_forms = set()

    for key, cfg in list(db.get("components_by_model", {}).items()):
        if cfg.get("model_key") == model_key or normalize_key(key) == model_key:
            removed_sources.append(clean_text(cfg.get("source_file")) or key)
            removed_forms.update(clean_text(form) for form in cfg.get("forms", {}) if clean_text(form))
            del db["components_by_model"][key]

    for key, cfg in list(db.get("config_exports_by_model", {}).items()):
        if cfg.get("model_key") == model_key or normalize_key(key) == model_key:
            removed_sources.append(clean_text(cfg.get("source_file")) or key)
            removed_forms.update(clean_text(form) for form in cfg.get("forms", {}) if clean_text(form))
            del db["config_exports_by_model"][key]

    if removed_sources:
        deleted_keys = {normalize_key(key) for key in db.get("deleted_config_model_keys", []) if clean_text(key)}
        deleted_keys.add(model_key)
        db["deleted_config_model_keys"] = sorted(deleted_keys)
        published_keys = {normalize_key(key) for key in db.get("published_config_model_keys", []) if clean_text(key)}
        published_keys.discard(model_key)
        db["published_config_model_keys"] = sorted(published_keys)
        rejected_keys = {normalize_key(key) for key in db.get("rejected_config_model_keys", []) if clean_text(key)}
        rejected_keys.discard(model_key)
        db["rejected_config_model_keys"] = sorted(rejected_keys)
        save_database(db)
    return sorted(set(removed_sources)), sorted(removed_forms)


def is_config_published(db, model_name):
    model_key = normalize_key(model_name)
    return model_key in {normalize_key(key) for key in db.get("published_config_model_keys", []) if clean_text(key)}


def is_config_rejected(db, model_name):
    model_key = normalize_key(model_name)
    return model_key in {normalize_key(key) for key in db.get("rejected_config_model_keys", []) if clean_text(key)}


def publish_tower_config(model_name):
    db = load_database()
    if not has_imported_config(db, model_name):
        raise ValueError(f"{model_name} 尚未导入配置表，不能发布。")
    if is_config_rejected(db, model_name):
        raise ValueError(f"{model_name} 已被拒绝发布，请重新导入配置表后再审核发布。")
    model_key = normalize_key(model_name)
    published_keys = {normalize_key(key) for key in db.get("published_config_model_keys", []) if clean_text(key)}
    was_pending = model_key not in published_keys
    published_keys.add(model_key)
    db["published_config_model_keys"] = sorted(published_keys)
    rejected_keys = {normalize_key(key) for key in db.get("rejected_config_model_keys", []) if clean_text(key)}
    rejected_keys.discard(model_key)
    db["rejected_config_model_keys"] = sorted(rejected_keys)
    save_database(db)
    return was_pending


def reject_tower_config(model_name):
    db = load_database()
    if not has_imported_config(db, model_name):
        raise ValueError(f"{model_name} 尚未导入配置表，不能拒绝发布。")
    model_key = normalize_key(model_name)
    published_keys = {normalize_key(key) for key in db.get("published_config_model_keys", []) if clean_text(key)}
    published_keys.discard(model_key)
    rejected_keys = {normalize_key(key) for key in db.get("rejected_config_model_keys", []) if clean_text(key)}
    rejected_keys.add(model_key)
    db["published_config_model_keys"] = sorted(published_keys)
    db["rejected_config_model_keys"] = sorted(rejected_keys)
    save_database(db)


def register_fonts():
    for name, path in [
        ("AppFont", r"C:\Windows\Fonts\ARIALUNI.ttf"),
        ("AppFont", r"C:\Windows\Fonts\msyh.ttc"),
        ("AppFont", r"C:\Windows\Fonts\simhei.ttf"),
    ]:
        if os.path.exists(path):
            try:
                pdfmetrics.registerFont(TTFont(name, path))
                return name
            except Exception:
                continue
    return "Helvetica"


def embedded_logo_image(width, height):
    return Image(BytesIO(base64.b64decode(EMBEDDED_LOGO_PNG)), width=width, height=height)


def make_paragraph(text, style):
    escaped = clean_text(text).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
    return Paragraph(escaped.replace("\n", "<br/>"), style)


def label(zh, en=None, lang="zh"):
    if lang == "zh":
        return zh
    translated = TRANSLATIONS.get(zh, {}).get(lang)
    if translated:
        return translated
    return en or TRANSLATIONS.get(zh, {}).get("en") or zh


def tr_text(text, lang, fallback=None):
    text = clean_text(text)
    if not text or lang == "zh":
        return text
    if text in TRANSLATIONS:
        return TRANSLATIONS[text].get(lang) or TRANSLATIONS[text].get("en") or text
    if "\n" in text:
        return "\n".join(tr_text(part, lang) for part in text.split("\n"))
    if "\\" in text:
        return "\\".join(tr_text(part, lang) for part in text.split("\\"))
    replaced = text
    for zh in sorted(TRANSLATIONS, key=len, reverse=True):
        if zh in replaced:
            replacement = TRANSLATIONS[zh].get(lang) or TRANSLATIONS[zh].get("en") or zh
            def contextual_replacement(match, value=replacement):
                prefix = " " if match.start() > 0 and re.search(r"[0-9A-Za-z)]$", replaced[:match.start()]) and re.search(r"^[0-9A-Za-z]", value) else ""
                suffix = " " if match.end() < len(replaced) and re.search(r"[0-9A-Za-z]$", value) and re.search(r"^[0-9A-Za-z(]", replaced[match.end():]) else ""
                return f"{prefix}{value}{suffix}"

            replaced = re.sub(re.escape(zh), contextual_replacement, replaced)
    if replaced != text:
        return remove_remaining_cjk(replaced, lang, fallback)
    if "-" in text:
        parts = text.split("-")
        converted = [tr_text(part.strip(), lang) for part in parts]
        if converted != parts:
            return remove_remaining_cjk(" - ".join(converted), lang, fallback)
    return remove_remaining_cjk(fallback or text, lang)


def remove_remaining_cjk(text, lang, fallback=None):
    text = clean_text(text)
    fallback = clean_text(fallback)
    if not re.search(r"[\u4e00-\u9fff]", text):
        return text
    if fallback and not re.search(r"[\u4e00-\u9fff]", fallback):
        return fallback
    unknown = {
        "en": "Untranslated item",
        "fr": "Element non traduit",
        "de": "Nicht uebersetzter Eintrag",
    }.get(lang, "Untranslated item")
    text = re.sub(r"[\u4e00-\u9fff]+", unknown, text)
    text = re.sub(rf"(?:{re.escape(unknown)}[\s:：,，;；()（）-]*)+", unknown, text)
    return text.strip(" :：,，;；-()（）") or unknown


def find_export_config(db, model_name):
    model_key = normalize_key(model_name)
    for cfg in db.get("config_exports_by_model", {}).values():
        if cfg.get("model_key") == model_key:
            return cfg
    return {}


def find_component_config(db, model_name):
    model_key = normalize_key(model_name)
    for cfg in db.get("components_by_model", {}).values():
        if cfg.get("model_key") == model_key:
            return cfg
    return {}


def is_config_deleted(db, model_name):
    model_key = normalize_key(model_name)
    return any(normalize_key(key) == model_key for key in db.get("deleted_config_model_keys", []))


def has_imported_config(db, model_name):
    if is_config_deleted(db, model_name):
        return False
    return bool(find_component_config(db, model_name) or find_export_config(db, model_name))


def find_initial_export_config(model_name):
    if DB_FILE.exists() or not INITIAL_DB_FILE.exists() or INITIAL_DB_FILE == DB_FILE:
        return {}
    try:
        with INITIAL_DB_FILE.open("r", encoding="utf-8") as f:
            initial_db = json.load(f)
    except Exception:
        return {}
    return find_export_config(initial_db, model_name)


def find_source_workbook_path(source_file):
    source_file = clean_text(source_file)
    if not source_file:
        return None
    direct = Path(source_file)
    if direct.exists():
        return direct
    search_roots = []
    for root in [
        APP_DIR,
        APP_DIR.parent,
        BUNDLE_DIR,
        DATA_DIR,
        Path.cwd(),
        Path.home() / "Documents" / "中联塔机报价单程序",
    ]:
        for candidate in [root, root / "参考文件", root / "data"]:
            if candidate.exists() and candidate not in search_roots:
                search_roots.append(candidate)
    for root in search_roots:
        try:
            matches = list(root.rglob(source_file)) if root.is_dir() else []
        except Exception:
            matches = []
        if matches:
            return matches[0]
    return None


def rebuild_export_config_from_workbook_bytes(db, model_name, workbook_bytes, source_file):
    wb = load_workbook(BytesIO(workbook_bytes), data_only=True)
    ws = wb.active
    form_row, form_columns = find_config_form_columns(ws)
    export_forms = build_config_export_rows(ws, form_row, form_columns)
    model_key = normalize_key(model_name)
    db.setdefault("config_exports_by_model", {})[model_name] = {
        "model": model_name,
        "model_key": model_key,
        "source_file": clean_text(source_file),
        "workbook_base64": base64.b64encode(workbook_bytes).decode("ascii"),
        "export_schema_version": EXPORT_SCHEMA_VERSION,
        "forms": export_forms,
    }
    save_database(db)
    return load_database()


def ensure_export_config(db, model_name):
    cfg = find_export_config(db, model_name)
    if cfg and int(cfg.get("export_schema_version", 0) or 0) >= EXPORT_SCHEMA_VERSION:
        return db, cfg

    if cfg:
        workbook_base64 = clean_text(cfg.get("workbook_base64", ""))
        if workbook_base64:
            try:
                refreshed_db = rebuild_export_config_from_workbook_bytes(
                    db,
                    model_name,
                    base64.b64decode(workbook_base64),
                    cfg.get("source_file", ""),
                )
                rebuilt_cfg = find_export_config(refreshed_db, model_name)
                if rebuilt_cfg:
                    return refreshed_db, rebuilt_cfg
            except Exception:
                pass

    initial_cfg = find_initial_export_config(model_name)
    if initial_cfg and int(initial_cfg.get("export_schema_version", 0) or 0) >= EXPORT_SCHEMA_VERSION:
        db.setdefault("config_exports_by_model", {})[model_name] = initial_cfg
        save_database(db)
        refreshed_db = load_database()
        cfg = find_export_config(refreshed_db, model_name)
        if cfg:
            return refreshed_db, cfg

    component_cfg = find_component_config(db, model_name)
    workbook_base64 = clean_text(component_cfg.get("workbook_base64", ""))
    if workbook_base64:
        try:
            refreshed_db = rebuild_export_config_from_workbook_bytes(
                db,
                model_name,
                base64.b64decode(workbook_base64),
                component_cfg.get("source_file", ""),
            )
            cfg = find_export_config(refreshed_db, model_name)
            if cfg:
                return refreshed_db, cfg
        except Exception:
            pass

    source_path = find_source_workbook_path(component_cfg.get("source_file", ""))
    if not source_path:
        raise ValueError("当前机型已导入配置表摘要，但旧版本没有保存原 Excel 内容，且找不到原配置表文件，无法自动生成清单数据。请把原配置表文件放回软件目录或参考文件目录。")

    import_tower_config_file(source_path, model_name=model_name)
    refreshed_db = load_database()
    cfg = find_export_config(refreshed_db, model_name)
    if not cfg:
        raise ValueError("已找到原配置表，但自动生成清单数据失败。")
    return refreshed_db, cfg


def make_config_list_excel(db, model_name, install_form, list_type, output_path, lang="zh"):
    try:
        db, cfg = ensure_export_config(db, model_name)
    except ValueError:
        if list_type == "basic":
            return make_basic_list_from_component_summary(db, model_name, install_form, output_path, lang=lang)
        raise
    form_key = normalize_form(install_form)
    form_data = cfg.get("forms", {}).get(form_key)
    if not form_data:
        for key, value in cfg.get("forms", {}).items():
            normalized_key = normalize_form(key)
            if normalized_key in form_key or form_key in normalized_key:
                form_data = value
                break
    if not form_data:
        available = "、".join(cfg.get("forms", {}).keys()) or "无"
        raise ValueError(f"当前安装形式没有可导出的配置表数据。当前选择：{install_form}；配置表中可用安装形式：{available}")

    row_key = "basic_rows" if list_type == "basic" else "option_rows"
    rows = form_data.get(row_key, [])
    if not rows:
        raise ValueError("没有符合条件的行项目可导出。")

    title = tr_text("基本配置清单" if list_type == "basic" else "增减配清单", lang)
    wb = Workbook()
    ws = wb.active
    ws.title = title
    header_rows = form_data.get("headers", [])
    if list_type == "basic":
        header_rows = header_rows[2:]
        display_header_rows = []
        for idx, row in enumerate(header_rows):
            display_header_rows.append((["序号"] if idx == 0 else [""]) + row[1:])
        display_rows = [[idx] + row[1:] for idx, row in enumerate(rows, start=1)]
    else:
        display_header_rows = header_rows
        display_rows = rows
    max_cols = max([len(row) for row in display_header_rows + display_rows] or [1])

    display_header_rows = [[tr_text(value, lang) if isinstance(value, str) else value for value in row] for row in display_header_rows]
    display_rows = [[tr_text(value, lang) if isinstance(value, str) else value for value in row] for row in display_rows]

    ws.append([title])
    if list_type == "basic":
        ws.append([f"{label('产品型号', 'Model', lang)}: {model_name}", f"{label('安装形式', 'Installation form', lang)}: {tr_text(install_form, lang)}"])
    else:
        ws.append([
            f"{label('产品型号', 'Model', lang)}: {model_name}",
            f"{label('安装形式', 'Installation form', lang)}: {tr_text(install_form, lang)}",
            f"{tr_text('来源配置表：', lang)}{cfg.get('source_file', '')}",
        ])
    ws.append([])
    for row in display_header_rows:
        ws.append(row)
    for row in display_rows:
        ws.append(row)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_cols)
    merge_key = "basic_merges" if list_type == "basic" else "option_merges"
    for merged in form_data.get(merge_key, []):
        if list_type == "basic":
            if int(merged.get("min_row", 0)) <= 2:
                continue
            if int(merged.get("max_row", 0)) <= 4:
                continue
            if int(merged.get("min_col", 0)) <= 1:
                continue
            min_row = int(merged.get("min_row", 0)) + 1
            max_row = int(merged.get("max_row", 0)) + 1
            min_col = int(merged.get("min_col", 0))
            max_col = int(merged.get("max_col", 0))
        else:
            min_row = int(merged.get("min_row", 0)) + 3
            max_row = int(merged.get("max_row", 0)) + 3
            min_col = int(merged.get("min_col", 0))
            max_col = int(merged.get("max_col", 0))
        if min_row and max_row and min_col and max_col and (min_row != max_row or min_col != max_col):
            ws.merge_cells(start_row=min_row, start_column=min_col, end_row=max_row, end_column=max_col)
    if list_type == "basic" and ws.max_row >= 5:
        for col_idx in range(1, max_cols + 1):
            top_value = clean_text(ws.cell(4, col_idx).value)
            bottom_value = clean_text(ws.cell(5, col_idx).value)
            ws.cell(4, col_idx).value = bottom_value or top_value
            ws.cell(5, col_idx).value = ""
            ws.merge_cells(start_row=4, start_column=col_idx, end_row=5, end_column=col_idx)
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    for col_idx in range(1, max_cols + 1):
        width = 12
        for row_idx in range(1, ws.max_row + 1):
            value = clean_text(ws.cell(row_idx, col_idx).value)
            if value:
                width = max(width, min(40, len(value) + 2))
        ws.column_dimensions[column_label(col_idx)].width = width
    ws.freeze_panes = "A4"
    return save_workbook_without_overwrite_conflict(wb, output_path)


def make_basic_list_from_component_summary(db, model_name, install_form, output_path, lang="zh"):
    components = select_components(db, model_name, install_form)
    if not components:
        raise ValueError("当前机型已导入配置表摘要，但没有可生成基本配置清单的主要部件数据。")

    wb = Workbook()
    ws = wb.active
    ws.title = safe_sheet_title(tr_text("基本配置清单", lang))
    ws.append([tr_text("基本配置清单", lang)])
    ws.append([f"{label('产品型号', 'Model', lang)}: {model_name}", f"{label('安装形式', 'Installation form', lang)}: {tr_text(install_form, lang)}"])
    ws.append([])
    ws.append([label("序号", "No.", lang), label("部件", "Component", lang), label("名称", "Name", lang), label("代号", "Code", lang), label("数量", "Qty", lang), label("标配", "Standard", lang)])
    for idx, item in enumerate(components, start=1):
        ws.append([
            idx,
            tr_text(item.get("component_zh", ""), lang, item.get("component_en", "")),
            tr_text(item.get("part_name", ""), lang),
            tr_text(item.get("code", ""), lang),
            item.get("quantity", ""),
            "●",
        ])

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    for col_idx in range(1, 7):
        width = 12
        for row_idx in range(1, ws.max_row + 1):
            value = clean_text(ws.cell(row_idx, col_idx).value)
            if value:
                width = max(width, min(40, len(value) + 2))
        ws.column_dimensions[column_label(col_idx)].width = width
    return save_workbook_without_overwrite_conflict(wb, output_path)


def safe_sheet_title(title):
    text = re.sub(r"[:\\/?*\[\]]", "_", clean_text(title)) or "Sheet"
    return text[:31]


def safe_filename_stem(name):
    text = re.sub(r'[<>:"/\\|?*]+', "_", clean_text(name))
    text = re.sub(r"\s+", "_", text)
    return text.strip(" ._") or "配置及增减配清单"


def workbook_version_filename_suffix(db, model_name):
    workbook_bytes = workbook_bytes_for_model(db, model_name)
    if not workbook_bytes:
        return ""
    try:
        wb = load_workbook(BytesIO(workbook_bytes), data_only=True, read_only=True)
        value = clean_text(wb.worksheets[0].cell(2, 1).value)
    except Exception:
        return ""
    return safe_filename_stem(value)


def infer_model_from_config_filename(path, products):
    stem = Path(path).stem.replace("（", "(").replace("）", ")")
    candidates = [
        stem,
        re.sub(r"(配置表|配置及增减配清单|配置清单|增减配清单)$", "", stem).strip(),
    ]
    product_map = {normalize_key(model): model for model in products}
    for candidate in candidates:
        matched = product_map.get(normalize_key(candidate))
        if matched:
            return matched
    return ""


def export_all_form_column_indexes(ws, form_row, form_columns, drop_price_only=False, start_row=1, end_row=None):
    end_row = end_row or ws.max_row
    columns = []
    component_col = None
    name_col = None
    for col in range(1, ws.max_column + 1):
        normalized_header = re.sub(r"\s+", "", column_header_text(ws, col, form_row, start_row))
        if normalized_header == "部件":
            component_col = col
        elif normalized_header == "名称":
            name_col = col
    for col in range(1, ws.max_column + 1):
        header_text = column_header_text(ws, col, form_row, start_row)
        normalized_header = re.sub(r"\s+", "", header_text)
        if not drop_price_only and component_col and name_col and col == component_col:
            continue
        drop_keywords = ["编码", "价格", "基价"] if drop_price_only else [keyword for keyword in EXPORT_DROP_KEYWORDS if keyword != "序号"]
        if any(keyword in normalized_header for keyword in drop_keywords):
            continue
        has_data = any(clean_text(ws.cell(row_idx, col).value) for row_idx in range(start_row, end_row + 1))
        if header_text or has_data or col in form_columns.values():
            columns.append(col)
    return columns


def copy_cell_for_export(source_cell, target_cell, value=None):
    target_cell.value = source_cell.value if value is None else value
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.fill = copy(source_cell.fill)
        target_cell.border = copy(source_cell.border)
        target_cell.alignment = copy(source_cell.alignment)
        target_cell.number_format = source_cell.number_format
        target_cell.protection = copy(source_cell.protection)
    if source_cell.hyperlink:
        target_cell._hyperlink = copy(source_cell.hyperlink)
    if source_cell.comment:
        target_cell.comment = copy(source_cell.comment)


def is_gray_fill(cell):
    fill = cell.fill
    if not fill or fill.fill_type is None:
        return False
    color = fill.fgColor
    value = clean_text(color.rgb or color.indexed or color.theme)
    if not value:
        return False
    return value.upper() not in ("00000000", "FFFFFFFF", "0", "1")


def normalize_export_sheet_style(ws):
    for row in ws.iter_rows():
        for cell in row:
            should_bold = cell.row == 1 or is_gray_fill(cell)
            cell.font = Font(name="Arial", size=10, bold=should_bold)


def config_list_title(source_title, model_name, list_type):
    text = clean_text(source_title)
    if list_type == "basic":
        if "选配配置表" in text:
            return text.replace("选配配置表", "基础配置表")
        if "增减配" in text:
            return text.replace("增减配", "基础配置")
        return text or f"{model_name}塔机基础配置表"
    if "基础配置表" in text:
        return text.replace("基础配置表", "选配配置表")
    if "基本配置表" in text:
        return text.replace("基本配置表", "选配配置表")
    if "配置表" in text and "选配配置表" not in text:
        return text.replace("配置表", "选配配置表", 1)
    return text or f"{model_name}塔机选配配置表"


def selected_config_source_rows(ws, form_row, form_columns, list_type, start_row=1, end_row=None):
    end_row = end_row or ws.max_row
    source_rows = list(range(start_row, form_row + 1))
    for row_idx in range(form_row + 1, end_row + 1):
        full_row_text = " ".join(clean_text(ws.cell(row_idx, col).value) for col in range(1, ws.max_column + 1))
        if not full_row_text:
            continue
        if "图" in full_row_text or "电源箱总成" in full_row_text:
            continue
        marks = [clean_text(ws.cell(row_idx, col).value) for col in form_columns.values()]
        keyword_match = any(keyword in full_row_text for keyword in EXPORT_OPTION_KEYWORDS)
        if list_type == "basic" and any("●" in mark for mark in marks):
            source_rows.append(row_idx)
        elif list_type == "option" and (any("○" in mark for mark in marks) or keyword_match):
            source_rows.append(row_idx)
    return source_rows


def copy_config_sheet(source_ws, target_ws, source_rows, source_columns, form_row, form_columns, list_type, model_name, normalize_headers=True, renumber=True, lang="zh"):
    row_map = {source_row: target_row for target_row, source_row in enumerate(source_rows, start=1)}
    col_map = {source_col: target_col for target_col, source_col in enumerate(source_columns, start=1)}
    sequence_col = None
    for col in source_columns:
        if "序号" in column_header_text(source_ws, col, form_row, source_rows[0] if source_rows else 1):
            sequence_col = col
            break

    body_index = 1
    for target_row, source_row in enumerate(source_rows, start=1):
        if source_ws.row_dimensions[source_row].height:
            target_ws.row_dimensions[target_row].height = source_ws.row_dimensions[source_row].height
        for target_col, source_col in enumerate(source_columns, start=1):
            source_cell = source_ws.cell(source_row, source_col)
            value = source_cell.value
            if normalize_headers and source_row == 1 and source_col == source_columns[0]:
                value = config_list_title(value, model_name, list_type)
            elif normalize_headers and source_row <= form_row:
                normalized_value = re.sub(r"\s+", "", clean_text(value))
                if normalized_value == "组成":
                    value = "产品模块"
                elif normalized_value == "名称":
                    value = "部件名称"
            elif renumber and source_row > form_row and source_col == sequence_col:
                value = body_index
            elif list_type == "option" and source_row > form_row and source_col in form_columns.values():
                row_text = " ".join(clean_text(source_ws.cell(source_row, col).value) for col in range(1, source_ws.max_column + 1))
                if any(keyword in row_text for keyword in EXPORT_OPTION_KEYWORDS):
                    value = "○"
            if isinstance(value, str):
                value = tr_text(value, lang)
            copy_cell_for_export(source_cell, target_ws.cell(target_row, target_col), value=value)
        if source_row > form_row:
            body_index += 1

    for source_col, target_col in col_map.items():
        width = source_ws.column_dimensions[column_label(source_col)].width
        if width:
            target_ws.column_dimensions[column_label(target_col)].width = width

    for merged in source_ws.merged_cells.ranges:
        mapped_rows = sorted(
            row_map[row]
            for row in range(merged.min_row, merged.max_row + 1)
            if row in row_map
        )
        mapped_cols = sorted(
            col_map[col]
            for col in range(merged.min_col, merged.max_col + 1)
            if col in col_map
        )
        if not mapped_rows or not mapped_cols:
            continue
        rows_are_contiguous = mapped_rows == list(range(mapped_rows[0], mapped_rows[-1] + 1))
        cols_are_contiguous = mapped_cols == list(range(mapped_cols[0], mapped_cols[-1] + 1))
        if not rows_are_contiguous or not cols_are_contiguous:
            continue
        if mapped_rows[0] == mapped_rows[-1] and mapped_cols[0] == mapped_cols[-1]:
            continue
        target_ws.merge_cells(
            start_row=mapped_rows[0],
            start_column=mapped_cols[0],
            end_row=mapped_rows[-1],
            end_column=mapped_cols[-1],
        )

    target_ws.freeze_panes = source_ws.freeze_panes
    target_ws.sheet_view.showGridLines = source_ws.sheet_view.showGridLines
    normalize_export_sheet_style(target_ws)


def make_combined_config_option_excel(db, model_name, output_path, lang="zh"):
    workbook_bytes = workbook_bytes_for_model(db, model_name)
    if not workbook_bytes:
        raise ValueError("当前机型没有已导入的配置表原始数据，无法生成配置及增减配清单。")

    source_wb = load_workbook(BytesIO(workbook_bytes))
    basic_source_ws = worksheet_for_list_type(source_wb, "basic")
    option_source_ws = worksheet_for_list_type(source_wb, "option")
    basic_start, basic_end = list_section_bounds(basic_source_ws, "basic")
    option_start, option_end = list_section_bounds(option_source_ws, "option")
    basic_form_row, basic_form_columns = find_config_form_columns(basic_source_ws, basic_start, basic_end)
    option_form_row, option_form_columns = find_config_form_columns(option_source_ws, option_start, option_end)
    if not basic_form_columns:
        raise ValueError("配置表中未识别到安装形式列，无法生成清单。")
    if not option_form_columns:
        option_source_ws = basic_source_ws
        option_start, option_end = basic_start, basic_end
        option_form_row, option_form_columns = basic_form_row, basic_form_columns

    basic_columns = export_all_form_column_indexes(basic_source_ws, basic_form_row, basic_form_columns, drop_price_only=True, start_row=basic_start, end_row=basic_end)
    option_columns = export_all_form_column_indexes(option_source_ws, option_form_row, option_form_columns, drop_price_only=True, start_row=option_start, end_row=option_end)
    if not basic_columns or not option_columns:
        raise ValueError("配置表中没有可导出的列。")

    output_wb = Workbook()
    basic_ws = output_wb.active
    basic_ws.title = safe_sheet_title(f"{model_name}{tr_text('配置清单', lang)}")
    option_ws = output_wb.create_sheet(safe_sheet_title(f"{model_name}{tr_text('增减配清单', lang)}"))

    basic_rows = selected_config_source_rows(basic_source_ws, basic_form_row, basic_form_columns, "basic", basic_start, basic_end)
    option_rows = selected_config_source_rows(option_source_ws, option_form_row, option_form_columns, "option", option_start, option_end)
    if len(basic_rows) <= (basic_form_row - basic_start + 1):
        raise ValueError("配置表中没有识别到基础配置实心圈项目。")
    if len(option_rows) <= (option_form_row - option_start + 1):
        raise ValueError("配置表中没有识别到增减配项目。")

    copy_config_sheet(
        basic_source_ws,
        basic_ws,
        basic_rows,
        basic_columns,
        basic_form_row,
        basic_form_columns,
        "basic",
        model_name,
        normalize_headers=False,
        renumber=False,
        lang=lang,
    )
    copy_config_sheet(
        option_source_ws,
        option_ws,
        option_rows,
        option_columns,
        option_form_row,
        option_form_columns,
        "option",
        model_name,
        normalize_headers=False,
        renumber=False,
        lang=lang,
    )

    return save_workbook_without_overwrite_conflict(output_wb, output_path)


def tower_type_text(tower_type, lang):
    if lang == "zh":
        return tower_type.get("zh", "")
    if lang == "en":
        return tower_type.get("en", "") or tr_text(tower_type.get("zh", ""), lang)
    return tr_text(tower_type.get("zh", ""), lang, tower_type.get("en", ""))


def mm_value(value, suffix="m"):
    text = clean_text(value)
    if not text:
        return ""
    return text if re.search(r"[a-zA-Z米mM]$", text) else f"{text}{suffix}"


def select_components(db, model_name, install_form):
    if not is_config_published(db, model_name):
        return []

    def visible(items):
        return [
            item
            for item in items
            if not should_skip_component(" ".join([item.get("component_zh", ""), item.get("part_name", ""), item.get("code", "")]))
        ]

    model_key = normalize_key(model_name)
    form_key = normalize_form(install_form)
    for cfg in db.get("components_by_model", {}).values():
        if cfg.get("model_key") != model_key:
            continue
        forms = cfg.get("forms", {})
        if form_key in forms:
            return visible(forms[form_key])
        for saved_form, items in forms.items():
            if form_key in saved_form or saved_form in form_key:
                return visible(items)
    return []


def change_type_text(option, lang):
    code = clean_text(option.get("change_type_code"))
    if code == "20":
        return label("减配", "Deduction", lang)
    return label("增配", "Addition", lang)


@dataclass
class QuoteInput:
    language: str
    model_name: str
    install_form: str
    quote_date: str
    trade_term: str
    trade_place: str
    price: str
    currency: str
    header_address: str
    customer_name: str
    quote_company: str
    quote_person: str
    quote_phone: str
    quote_email: str
    payment: str
    delivery: str
    validity: str
    transportation: str
    warranty: str
    others: str
    selected_option_indexes: list
    selected_option_quantities: dict
    selected_option_items: list


def make_pdf(db, quote, output_path):
    ensure_dirs()
    lang = quote.language
    font_name = register_fonts()
    styles = getSampleStyleSheet()
    normal = ParagraphStyle("normal", parent=styles["Normal"], fontName=font_name, fontSize=7.6, leading=9.2)
    small = ParagraphStyle("small", parent=normal, fontSize=6.6, leading=7.8)
    address_style = ParagraphStyle("address", parent=normal, fontSize=7.4, leading=8.8)
    title = ParagraphStyle("title", parent=normal, alignment=TA_CENTER, fontSize=15, leading=18)
    section = ParagraphStyle("section", parent=normal, fontSize=8.6, leading=10, alignment=TA_LEFT)
    center = ParagraphStyle("center", parent=normal, alignment=TA_CENTER)
    left = ParagraphStyle("left", parent=normal, alignment=TA_LEFT)

    doc = BaseDocTemplate(
        str(output_path),
        pagesize=A4,
        leftMargin=17 * mm,
        rightMargin=17 * mm,
        topMargin=12 * mm,
        bottomMargin=10 * mm,
    )
    frame = Frame(doc.leftMargin, doc.bottomMargin, doc.width, doc.height, id="normal")
    doc.addPageTemplates([PageTemplate(id="quotation", frames=[frame])])
    width = doc.width

    product = db["products"].get(quote.model_name, {})
    product_code = product.get("product_code", "")
    forms = db.get("forms_by_code", {}).get(product_code, [])
    form = next((item for item in forms if normalize_form(item.get("install_form")) == normalize_form(quote.install_form)), {})
    tower_type = product.get("tower_type") or detect_tower_type(quote.model_name)
    max_load = product.get("max_load") or detect_max_load(quote.model_name)
    selected_options = quote.selected_option_items or []
    components = select_components(db, quote.model_name, quote.install_form)

    story = []
    company_address = quote.header_address
    logo = embedded_logo_image(width=36 * mm, height=11 * mm)
    header = Table(
        [
            [logo, make_paragraph(company_address, address_style)],
            [make_paragraph(label("中联塔机报价单", "TOWER CRANE QUOTATION", lang), title), ""],
            [make_paragraph(f"{label('报价日期', 'Date', lang)}: {quote.quote_date}", section), ""],
        ],
        colWidths=[39 * mm, width - 39 * mm],
    )
    header.setStyle(
        TableStyle(
            [
                ("GRID", (0, 0), (-1, -1), 0.8, colors.black),
                ("SPAN", (0, 1), (1, 1)),
                ("SPAN", (0, 2), (1, 2)),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("BACKGROUND", (0, 1), (1, 1), colors.whitesmoke),
            ]
        )
    )
    story.append(header)

    quote_info_rows = [
        [
            make_paragraph(label("客户名称", "Customer", lang), center),
            make_paragraph(tr_text(quote.customer_name, lang), left),
            make_paragraph(label("报价日期", "Date", lang), center),
            make_paragraph(quote.quote_date, left),
        ],
        [
            make_paragraph(label("报价单位", "Quotation unit", lang), center),
            make_paragraph(tr_text(quote.quote_company, lang), left),
            make_paragraph(label("报价人", "Quoted by", lang), center),
            make_paragraph(quote.quote_person, left),
        ],
        [
            make_paragraph(label("联系邮箱", "Email", lang), center),
            make_paragraph(quote.quote_email, left),
            make_paragraph(label("联系电话", "Phone", lang), center),
            make_paragraph(quote.quote_phone, left),
        ],
    ]
    quote_info = Table(quote_info_rows, colWidths=[26 * mm, width - 102 * mm, 26 * mm, 50 * mm])
    quote_info.setStyle(table_style())
    story.append(quote_info)

    unit_price = f"{quote.trade_term} {tr_text(quote.trade_place, lang)} ({label('价格', 'Price', lang)}: {quote.currency} {format_number(quote.price)})"
    basic_rows = [
        [make_paragraph(label("基本参数", "BASIC PARAMETER", lang), section), "", "", ""],
        [make_paragraph(label("产品型号", "Model", lang), center), make_paragraph(quote.model_name, center), make_paragraph(label("塔机类型", "Type", lang), center), make_paragraph(tower_type_text(tower_type, lang), center)],
        [make_paragraph(label("独立高度", "HUH", lang), center), make_paragraph(mm_value(form.get("height")), center), make_paragraph(label("最大臂长", "Max jib length", lang), center), make_paragraph(mm_value(form.get("jib_length") or product.get("default_jib_length")), center)],
        [make_paragraph(label("最大起重量", "Max load", lang), center), make_paragraph(max_load, center), make_paragraph(label("安装形式", "Installation form", lang), center), make_paragraph(tr_text(quote.install_form, lang), center)],
        [make_paragraph(label("容绳量", "Rope capacity", lang), center), make_paragraph(mm_value(form.get("rope_capacity")), center), make_paragraph(label("塔身种类", "Mast type", lang), center), make_paragraph(tr_text(product.get("mast_type", ""), lang), center)],
        [make_paragraph(label("单价", "UNIT PRICE", lang), section), make_paragraph(unit_price, section), "", ""],
    ]
    basic = Table(basic_rows, colWidths=[38 * mm, 48 * mm, 38 * mm, width - 124 * mm])
    basic.setStyle(table_style([(0, 0, 3, 0), (1, 5, 3, 5)]))
    story.append(basic)

    component_rows = [[make_paragraph(label("主要部件配置", "MAIN COMPONENT CONFIGURATION", lang), section), "", "", ""]]
    component_rows.append(
        [
            make_paragraph("No.", center),
            make_paragraph(label("部件名称", "Component", lang), center),
            make_paragraph(label("代号", "Code", lang), center),
            make_paragraph(label("数量", "Qty", lang), center),
        ]
    )
    if components:
        for idx, item in enumerate(components, start=1):
            component_name = tr_text(item.get("component_zh", ""), lang, item.get("component_en", ""))
            part_name = tr_text(item.get("part_name", ""), lang)
            display = component_name if part_name in ("", component_name) else f"{component_name} - {part_name}"
            component_rows.append(
                [
                    clean_text(idx),
                    make_paragraph(display, left),
                    make_paragraph(tr_text(code_or_slash(item.get("code", "")), lang), left),
                    make_paragraph(item.get("quantity", "1"), center),
                ]
            )
    else:
        component_rows.append(["", make_paragraph(label("未导入该机型配置表。", "No configuration workbook imported for this model.", lang), left), "", ""])
    comp_table = Table(component_rows, colWidths=[10 * mm, width - 80 * mm, 50 * mm, 20 * mm], repeatRows=2)
    comp_table.setStyle(table_style([(0, 0, 3, 0), (1, len(component_rows) - 1, 3, len(component_rows) - 1)] if not components else [(0, 0, 3, 0)]))
    story.append(comp_table)

    option_rows = [[make_paragraph(label("增减配置", "ADDITIONS / DEDUCTIONS", lang), section), "", "", "", ""]]
    option_rows.append(
        [
            make_paragraph("No.", center),
            make_paragraph(label("类型", "Type", lang), center),
            make_paragraph(label("部件代号", "Part code", lang), center),
            make_paragraph(label("增减部件", "Part", lang), center),
            make_paragraph(label("数量", "Qty", lang), center),
        ]
    )
    if selected_options:
        for idx, item in enumerate(selected_options, start=1):
            qty = clean_text(item.get("quantity", "")) or "1"
            option_rows.append(
                [
                    clean_text(idx),
                    make_paragraph(tr_text(item.get("change_type", "增配"), lang), center),
                    make_paragraph(tr_text(code_or_slash(item.get("model_code", "") or item.get("code", "")), lang), left),
                    make_paragraph(tr_text(item.get("item_display", "") or item.get("name", ""), lang), left),
                    make_paragraph(qty, center),
                ]
            )
    else:
        option_rows.append(["", make_paragraph(label("未选择增减配置。", "No additions or deductions selected.", lang), left), "", "", ""])
    option_table = Table(option_rows, colWidths=[10 * mm, 26 * mm, 42 * mm, width - 92 * mm, 14 * mm], repeatRows=2)
    option_table.setStyle(table_style([(0, 0, 4, 0), (1, len(option_rows) - 1, 4, len(option_rows) - 1)] if not selected_options else [(0, 0, 4, 0)]))
    story.append(option_table)

    terms_rows = [
        [make_paragraph(label("交易条款及其他信息", "TRADE CLAUSE & OTHER INFORMATION", lang), section), ""],
        [make_paragraph(label("付款方式", "Payment item", lang), center), make_paragraph(tr_text(quote.payment, lang), left)],
        [make_paragraph(label("交货期", "Delivery time", lang), center), make_paragraph(tr_text(quote.delivery, lang), left)],
        [make_paragraph(label("报价有效期", "Quotation validity", lang), center), make_paragraph(tr_text(quote.validity, lang), left)],
        [make_paragraph(label("运输方案", "Transportation", lang), center), make_paragraph(tr_text(quote.transportation, lang), left)],
        [make_paragraph(label("质保期", "Warranty", lang), center), make_paragraph(tr_text(quote.warranty, lang), small)],
        [make_paragraph(label("其他", "Others", lang), center), make_paragraph(tr_text(quote.others, lang), left)],
    ]
    terms = Table(terms_rows, colWidths=[40 * mm, width - 40 * mm])
    terms.setStyle(table_style([(0, 0, 1, 0)]))
    story.append(terms)

    doc.build(story)


def table_style(spans=None):
    commands = [
        ("GRID", (0, 0), (-1, -1), 0.55, colors.black),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 1),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 1),
        ("BACKGROUND", (0, 0), (-1, 0), colors.whitesmoke),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
    ]
    for c1, r1, c2, r2 in spans or []:
        commands.append(("SPAN", (c1, r1), (c2, r2)))
    return TableStyle(commands)


def make_ltc_option_pdf(quote, output_path):
    ensure_dirs()
    lang = quote.language
    font_name = register_fonts()
    styles = getSampleStyleSheet()
    normal = ParagraphStyle("normal", parent=styles["Normal"], fontName=font_name, fontSize=8, leading=9.5)
    title_style = ParagraphStyle("title", parent=normal, alignment=TA_CENTER, fontSize=16, leading=20)
    center = ParagraphStyle("center", parent=normal, alignment=TA_CENTER)
    left = ParagraphStyle("left", parent=normal, alignment=TA_LEFT)

    doc = BaseDocTemplate(
        str(output_path),
        pagesize=A4,
        leftMargin=14 * mm,
        rightMargin=14 * mm,
        topMargin=12 * mm,
        bottomMargin=10 * mm,
    )
    frame = Frame(doc.leftMargin, doc.bottomMargin, doc.width, doc.height, id="normal")
    doc.addPageTemplates([PageTemplate(id="ltc", frames=[frame])])
    story = [
        Paragraph(tr_text("LTC选配指导文件", lang), title_style),
        Paragraph(
            f"{label('产品型号', 'Model', lang)}: {quote.model_name}  "
            f"{label('安装形式', 'Installation form', lang)}: {tr_text(quote.install_form, lang)}  "
            f"{label('日期', 'Date', lang)}: {quote.quote_date}",
            normal,
        ),
    ]

    def ltc_text(value):
        return tr_text(clean_text(value).replace("图", ""), lang)

    rows = [[
        make_paragraph("No.", center),
        make_paragraph(label("增减配置", "Addition / deduction", lang), center),
        make_paragraph(label("部件", "Component", lang), center),
        make_paragraph(label("名称", "Name", lang), center),
        make_paragraph(label("编码", "Item No.", lang), center),
        make_paragraph(label("代号", "Code", lang), center),
        make_paragraph(label("数量", "Qty", lang), center),
    ]]
    expanded_items = []
    for item in quote.selected_option_items or []:
        children = item.get("children") or []
        if children:
            for child in children:
                expanded = dict(child)
                expanded["change_type"] = item.get("change_type", "增配")
                expanded["quantity"] = clean_text(item.get("quantity", "")) or clean_text(child.get("quantity", "")) or "1"
                expanded["package_name"] = item.get("item_display", "") or item.get("name", "")
                expanded_items.append(expanded)
        else:
            expanded_items.append(item)
    for idx, item in enumerate(expanded_items, start=1):
        rows.append([
            clean_text(idx),
            make_paragraph(ltc_text(item.get("change_type", "增配")), center),
            make_paragraph(ltc_text(item.get("component", "")), left),
            make_paragraph(ltc_text(item.get("name", "")), left),
            make_paragraph(ltc_text(item.get("code", "")), left),
            make_paragraph(ltc_text(code_or_slash(item.get("model_code", ""))), left),
            make_paragraph(clean_text(item.get("quantity", "")) or "1", center),
        ])
    table = Table(rows, colWidths=[8 * mm, 16 * mm, 31 * mm, 40 * mm, 34 * mm, 31 * mm, 13 * mm], repeatRows=1)
    table.setStyle(table_style())
    story.append(table)
    doc.build(story)


class ScrollableFrame(ttk.Frame):
    def __init__(self, parent):
        super().__init__(parent)
        self.canvas = Canvas(self, highlightthickness=0)
        self.scrollbar = ttk.Scrollbar(self, orient=VERTICAL, command=self.canvas.yview)
        self.inner = ttk.Frame(self.canvas)
        self.inner.bind("<Configure>", lambda _e: self.canvas.configure(scrollregion=self.canvas.bbox("all")))
        self.window_id = self.canvas.create_window((0, 0), window=self.inner, anchor="nw")
        self.canvas.bind("<Configure>", lambda e: self.canvas.itemconfigure(self.window_id, width=e.width))
        self.canvas.bind("<MouseWheel>", self.on_mousewheel)
        self.inner.bind("<MouseWheel>", self.on_mousewheel)
        self.canvas.configure(yscrollcommand=self.scrollbar.set)
        self.canvas.pack(side=LEFT, fill=BOTH, expand=True)
        self.scrollbar.pack(side=RIGHT, fill="y")

    def on_mousewheel(self, event):
        self.canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")


class QuotationApp:
    def __init__(self, root):
        self.root = root
        self.root.title(APP_TITLE)
        self.set_window_icon()
        self.root.geometry("1180x760")
        ensure_dirs()
        self.db = load_database()
        self.user_settings = load_user_settings()
        self.language_label = StringVar(value="中文")
        self.model_var = StringVar()
        self.model_display_var = StringVar()
        self.model_display_to_name = {}
        self.form_var = StringVar()
        self.date_var = StringVar(value=date.today().isoformat())
        self.trade_term = StringVar(value="FOB")
        self.trade_place = StringVar(value="")
        self.price = StringVar(value="")
        self.currency = StringVar(value="CNY")
        self.header_address = StringVar(value=self.user_settings["header_address"])
        self.customer_name = StringVar(value=self.user_settings["customer_name"])
        self.quote_company = StringVar(value=self.user_settings["quote_company"])
        self.quote_person = StringVar(value=self.user_settings["quote_person"])
        self.quote_phone = StringVar(value=self.user_settings["quote_phone"])
        self.quote_email = StringVar(value=self.user_settings["quote_email"])
        self.option_vars = []
        self.option_qty_vars = []
        self.option_type_vars = []
        self.config_option_items = []
        self.machine_price_var = StringVar(value="")
        self.option_price_var = StringVar(value="")
        self.total_price_var = StringVar(value="")
        self.term_defaults = {
            "payment": "合同签订后支付30%作为定金，发货前付清剩余70%合同款。",
            "delivery": "收到定金后90日内发货。",
            "validity": "30天",
            "transportation": "以最终发运方案为准。",
            "warranty": "自提单之日起保修期：钢结构12个月；机构12个月；电气部件12个月。易损件除外。",
            "others": "报价含首次安装指导服务费用。",
        }
        self.term_values = dict(self.term_defaults)
        self.term_widgets = {}
        self.quote_dialog = None
        self.research_operator_id = ""
        self.research_operation_time = ""
        self.header_icon_image = None
        self.build_ui()
        self.refresh_models()
        self.root.protocol("WM_DELETE_WINDOW", self.on_close)

    def set_window_icon(self):
        try:
            if WINDOW_ICON_FILE.exists():
                self.root.iconbitmap(default=str(WINDOW_ICON_FILE))
        except Exception:
            pass

    def build_ui(self):
        style = ttk.Style(self.root)
        style.configure("Title.TLabelframe.Label", font=("Microsoft YaHei UI", 11, "bold"))
        style.configure("DialogTitle.TLabel", font=("Microsoft YaHei UI", 12, "bold"))
        main = ttk.Frame(self.root, padding=12)
        main.pack(fill=BOTH, expand=True)
        title_row = ttk.Frame(main)
        title_row.pack(anchor="w", fill="x")
        try:
            if HEADER_ICON_FILE.exists():
                icon = PILImage.open(HEADER_ICON_FILE)
                icon.thumbnail((62, 62), PILImage.LANCZOS)
                self.header_icon_image = ImageTk.PhotoImage(icon)
                ttk.Label(title_row, image=self.header_icon_image).pack(side=LEFT, padx=(0, 10))
        except Exception:
            self.header_icon_image = None
        ttk.Label(title_row, text=APP_TITLE, font=("Microsoft YaHei UI", 18, "bold")).pack(side=LEFT)

        toolbar = ttk.Frame(main)
        toolbar.pack(fill="x", pady=(10, 8))
        ttk.Button(toolbar, text="配置及增减配清单", command=self.on_generate_combined_config_list).pack(side=LEFT, padx=(0, 8))
        ttk.Button(toolbar, text="报价单信息", command=self.open_quote_info_dialog).pack(side=LEFT, padx=(0, 8))
        ttk.Button(toolbar, text="生成报价 PDF", command=self.on_generate).pack(side=LEFT, padx=(0, 8))

        settings_row = ttk.Frame(main)
        settings_row.pack(fill="x")
        product_select = ttk.LabelFrame(settings_row, text="产品型号选择", padding=10, style="Title.TLabelframe")
        product_select.pack(side=LEFT, fill="both", expand=True, padx=(0, 8))
        info_box = ttk.LabelFrame(settings_row, text="标准配置信息", padding=6, style="Title.TLabelframe")
        info_box.pack(side=LEFT, fill="both", expand=True, padx=(0, 8))
        price_box = ttk.LabelFrame(settings_row, text="当前FOB参考价格", padding=8, style="Title.TLabelframe")
        price_box.pack(side=LEFT, fill="both")

        self.model_combo = ttk.Combobox(product_select, textvariable=self.model_display_var, width=28, state="readonly")
        self.form_combo = ttk.Combobox(product_select, textvariable=self.form_var, width=24, state="readonly")
        ttk.Label(product_select, text="产品型号").grid(row=0, column=0, sticky="w", padx=(0, 6), pady=4)
        self.model_combo.grid(row=0, column=1, sticky="we", pady=4)
        ttk.Label(product_select, text="安装形式").grid(row=1, column=0, sticky="w", padx=(0, 6), pady=4)
        self.form_combo.grid(row=1, column=1, sticky="we", pady=4)
        product_select.columnconfigure(1, weight=1)
        self.info_message_label = ttk.Label(info_box, text="", justify=LEFT)
        self.info_value_labels = {}
        info_items = [
            ("product_code", "产品编码"),
            ("tower_type", "塔机类型"),
            ("height", "独立高度"),
            ("jib_length", "最大臂长"),
            ("max_load", "最大起重量"),
            ("rope_capacity", "容绳量"),
            ("mast_type", "塔身种类"),
        ]
        for idx, (key, label_text) in enumerate(info_items):
            row = idx // 2
            col = (idx % 2) * 2
            ttk.Label(info_box, text=f"{label_text}:").grid(row=row, column=col, sticky="w", padx=(0, 3), pady=1)
            value_label = ttk.Label(info_box, text="", width=18)
            value_label.grid(row=row, column=col + 1, sticky="w", padx=(0, 8), pady=1)
            self.info_value_labels[key] = value_label
        self.info_message_label.grid(row=4, column=0, columnspan=4, sticky="w", pady=(3, 0))
        info_box.columnconfigure(1, weight=1)
        info_box.columnconfigure(3, weight=1)
        price_items = [
            ("整机价格", self.machine_price_var),
            ("选配价格", self.option_price_var),
            ("当前合计", self.total_price_var),
        ]
        for idx, (label_text, value_var) in enumerate(price_items):
            ttk.Label(price_box, text=f"{label_text}:").grid(row=idx, column=0, sticky="w", padx=(0, 6), pady=3)
            ttk.Label(price_box, textvariable=value_var, width=16).grid(row=idx, column=1, sticky="w", pady=3)
        self.model_combo.bind("<<ComboboxSelected>>", lambda _e: self.on_model_selected())
        self.form_combo.bind("<<ComboboxSelected>>", lambda _e: self.refresh_screen())

        body = ttk.PanedWindow(main, orient="horizontal")
        body.pack(fill=BOTH, expand=True, pady=10)
        left = ttk.Frame(body)
        right = ttk.Frame(body)
        body.add(left, weight=1)
        body.add(right, weight=2)

        basic_config_box = ttk.LabelFrame(left, text="产品基本配置", padding=10, style="Title.TLabelframe")
        basic_config_box.pack(fill=BOTH, expand=True)
        self.basic_config_tree = ttk.Treeview(
            basic_config_box,
            columns=("seq", "component", "name", "model_code", "mark"),
            show="headings",
            height=12,
        )
        basic_headings = {
            "seq": "序号",
            "component": "部件",
            "name": "名称",
            "model_code": "代号",
            "mark": "数量",
        }
        basic_widths = {"seq": 50, "component": 100, "name": 160, "model_code": 170, "mark": 80}
        for col in basic_headings:
            self.basic_config_tree.heading(col, text=basic_headings[col])
            self.basic_config_tree.column(col, width=basic_widths[col], anchor="w")
        basic_scroll = ttk.Scrollbar(basic_config_box, orient=VERTICAL, command=self.basic_config_tree.yview)
        self.basic_config_tree.configure(yscrollcommand=basic_scroll.set)
        self.basic_config_tree.pack(side=LEFT, fill=BOTH, expand=True)
        basic_scroll.pack(side=RIGHT, fill="y")

        option_box = ttk.LabelFrame(right, text="可选增减配置", padding=10, style="Title.TLabelframe")
        option_box.pack(fill=BOTH, expand=True)
        self.option_frame = ScrollableFrame(option_box)
        self.option_frame.pack(fill=BOTH, expand=True)

        bottom_bar = ttk.Frame(main)
        bottom_bar.pack(side="bottom", fill="x", anchor="w")
        ttk.Button(bottom_bar, text="配置及价格导入模块", command=self.open_research_import_module).pack(side=LEFT, padx=(0, 8))
        ttk.Button(bottom_bar, text="审核并发布", command=self.open_audit_publish_module).pack(side=LEFT, padx=(0, 8))
        ttk.Button(bottom_bar, text="后台数据变更记录", command=self.show_change_log).pack(side=LEFT, padx=(0, 8))

    def term_labels(self):
        return [
            ("payment", "付款方式"),
            ("delivery", "交货期"),
            ("validity", "报价有效期"),
            ("transportation", "运输方案"),
            ("warranty", "质保期"),
            ("others", "其他"),
        ]

    def sync_quote_dialog_values(self):
        live_widgets = {}
        for key, widget in self.term_widgets.items():
            if widget.winfo_exists():
                self.term_values[key] = widget.get("1.0", "end").strip()
                live_widgets[key] = widget
        self.term_widgets = live_widgets

    def open_quote_info_dialog(self):
        if self.quote_dialog is not None and self.quote_dialog.winfo_exists():
            self.quote_dialog.lift()
            self.quote_dialog.focus_force()
            return

        dialog = Toplevel(self.root)
        self.quote_dialog = dialog
        dialog.title("报价单信息")
        dialog.geometry("980x520")
        dialog.transient(self.root)

        wrapper = ttk.Frame(dialog, padding=12)
        wrapper.pack(fill=BOTH, expand=True)

        quote_info = ttk.LabelFrame(wrapper, text="报价单信息", padding=10, style="Title.TLabelframe")
        quote_info.pack(fill="x")
        self.language_combo = ttk.Combobox(quote_info, textvariable=self.language_label, values=list(LANG_OPTIONS), width=12, state="readonly")
        self.term_combo = ttk.Combobox(quote_info, textvariable=self.trade_term, values=["FOB", "CIF", "EXZ", "DDP", "DAP"], width=10, state="readonly")
        self.currency_combo = ttk.Combobox(quote_info, textvariable=self.currency, values=["CNY", "USD", "EUR"], width=10, state="readonly")
        quote_fields = [
            ("报价单语言", self.language_combo, 0, 0),
            ("报价日期", ttk.Entry(quote_info, textvariable=self.date_var, width=14), 0, 2),
            ("贸易术语", self.term_combo, 0, 4),
            ("交易地点", ttk.Entry(quote_info, textvariable=self.trade_place, width=18), 0, 6),
            ("价格", ttk.Entry(quote_info, textvariable=self.price, width=14), 1, 0),
            ("价格单位", self.currency_combo, 1, 2),
            ("报价单位", ttk.Entry(quote_info, textvariable=self.quote_company, width=38), 1, 4),
            ("客户名称", ttk.Entry(quote_info, textvariable=self.customer_name, width=38), 2, 0),
            ("报价人", ttk.Entry(quote_info, textvariable=self.quote_person, width=14), 3, 0),
            ("联系电话", ttk.Entry(quote_info, textvariable=self.quote_phone, width=18), 3, 2),
            ("联系邮箱", ttk.Entry(quote_info, textvariable=self.quote_email, width=30), 4, 0),
            ("公司地址", ttk.Entry(quote_info, textvariable=self.header_address, width=64), 5, 0),
        ]
        for label_text, widget, row, col in quote_fields:
            ttk.Label(quote_info, text=label_text).grid(row=row, column=col, sticky="w", padx=(0, 4), pady=3)
            span = 7 if label_text in ("公司地址", "联系邮箱", "客户名称") else (3 if label_text == "报价单位" else 1)
            widget.grid(row=row, column=col + 1, columnspan=span, sticky="we", padx=(0, 10), pady=3)
        quote_info.columnconfigure(7, weight=1)

        terms = ttk.LabelFrame(wrapper, text="交易条款及其他信息", padding=10, style="Title.TLabelframe")
        terms.pack(fill=BOTH, expand=True, pady=(10, 0))
        self.term_widgets = {}
        for idx, (key, text) in enumerate(self.term_labels()):
            row = idx // 2
            col = (idx % 2) * 2
            ttk.Label(terms, text=text).grid(row=row, column=col, sticky="nw", padx=(0, 4), pady=5)
            widget = Text(terms, width=42, height=3, wrap="word")
            widget.insert("1.0", self.term_values.get(key, self.term_defaults.get(key, "")))
            widget.grid(row=row, column=col + 1, sticky="nsew", padx=(0, 12), pady=5)
            self.term_widgets[key] = widget
        for col in (1, 3):
            terms.columnconfigure(col, weight=1)
        for row in range(3):
            terms.rowconfigure(row, weight=1)

        button_bar = ttk.Frame(wrapper)
        button_bar.pack(fill="x", pady=(10, 0))
        ttk.Button(button_bar, text="确定", command=lambda: self.close_quote_info_dialog(dialog)).pack(side=RIGHT)

        dialog.protocol("WM_DELETE_WINDOW", lambda: self.close_quote_info_dialog(dialog))

    def close_quote_info_dialog(self, dialog):
        self.sync_quote_dialog_values()
        self.save_current_user_settings()
        if dialog.winfo_exists():
            dialog.destroy()
        self.quote_dialog = None
        self.term_widgets = {}

    def current_product(self):
        return self.db.get("products", {}).get(self.model_var.get(), {})

    def current_forms(self):
        product_code = self.current_product().get("product_code", "")
        return self.db.get("forms_by_code", {}).get(product_code, [])

    def current_form(self):
        wanted = normalize_form(self.form_var.get())
        for form in self.current_forms():
            if normalize_form(form.get("install_form")) == wanted:
                return form
        return {}

    def current_options(self):
        product_code = self.current_product().get("product_code", "")
        return self.db.get("options_by_code", {}).get(product_code, [])

    def model_display_name(self, model):
        return model

    def on_model_selected(self):
        display = self.model_display_var.get()
        model = self.model_display_to_name.get(display, display)
        self.model_var.set(model)
        self.refresh_forms()

    def refresh_models(self):
        models = sorted(
            (model for model in self.db.get("products", {}) if is_visible_product_model(model)),
            key=lambda model: (0 if has_imported_config(self.db, model) else 1, model),
        )
        display_values = [self.model_display_name(model) for model in models]
        self.model_display_to_name = dict(zip(display_values, models))
        self.model_combo["values"] = display_values
        if models and self.model_var.get() not in models:
            self.model_var.set(models[0])
        elif not models:
            self.model_var.set("")
        self.model_display_var.set(self.model_display_name(self.model_var.get()) if self.model_var.get() else "")
        self.refresh_forms()

    def refresh_forms(self):
        forms = [item.get("install_form", "") for item in self.current_forms() if item.get("install_form")]
        self.form_combo["values"] = forms
        if forms and self.form_var.get() not in forms:
            self.form_var.set(forms[0])
        elif not forms:
            self.form_var.set("")
        self.refresh_screen()

    def refresh_screen(self):
        self.refresh_info()
        self.refresh_basic_config()
        self.refresh_options()

    def refresh_info(self):
        product = self.current_product()
        form = self.current_form()
        if not product:
            for label_widget in self.info_value_labels.values():
                label_widget.config(text="")
            self.info_message_label.config(text="请先导入增减配价格表，系统会从 D列=B1 的数据识别机型。")
            return
        type_info = product.get("tower_type", {})
        values = {
            "product_code": product.get("product_code", ""),
            "tower_type": type_info.get("zh", ""),
            "height": mm_value(form.get("height")),
            "jib_length": mm_value(form.get("jib_length") or product.get("default_jib_length")),
            "max_load": product.get("max_load", ""),
            "rope_capacity": mm_value(form.get("rope_capacity")),
            "mast_type": product.get("mast_type", ""),
        }
        for key, value in values.items():
            self.info_value_labels[key].config(text=value)
        self.info_message_label.config(text="")

    def refresh_options(self):
        for child in self.option_frame.inner.winfo_children():
            child.destroy()
        self.option_vars = []
        self.option_qty_vars = []
        self.option_type_vars = []
        self.config_option_items = []
        if not self.model_var.get() or not self.form_var.get():
            ttk.Label(self.option_frame.inner, text="请先选择产品型号和安装形式。").pack(anchor="w")
            self.update_current_price()
            return
        try:
            options = option_config_items(self.db, self.model_var.get(), self.form_var.get())
        except Exception as exc:
            ttk.Label(self.option_frame.inner, text=f"无法读取配置表增减配数据：{exc}").pack(anchor="w")
            self.update_current_price()
            return
        self.config_option_items = options
        if not options:
            ttk.Label(self.option_frame.inner, text="当前安装形式暂无可选增减配置。").pack(anchor="w")
            self.update_current_price()
            return

        header = ttk.Frame(self.option_frame.inner)
        header.pack(fill="x", pady=(0, 4))
        ttk.Label(header, text="选择", width=6).pack(side=LEFT)
        ttk.Label(header, text="数量", width=8).pack(side=LEFT, padx=(0, 6))
        ttk.Label(header, text="增减配", width=10).pack(side=LEFT, padx=(0, 6))
        ttk.Label(header, text="价格", width=12).pack(side=LEFT, padx=(0, 6))
        ttk.Label(header, text="项目").pack(side=LEFT)
        for option in options:
            row = ttk.Frame(self.option_frame.inner)
            row.pack(fill="x", anchor="w", pady=2)
            var = BooleanVar(value=False)
            qty_var = StringVar(value="1")
            type_var = StringVar(value="增配")
            var.trace_add("write", lambda *_args: self.update_current_price())
            qty_var.trace_add("write", lambda *_args: self.update_current_price())
            type_var.trace_add("write", lambda *_args: self.update_current_price())
            self.option_vars.append(var)
            self.option_qty_vars.append(qty_var)
            self.option_type_vars.append(type_var)
            check = ttk.Checkbutton(row, variable=var)
            qty = ttk.Spinbox(row, from_=1, to=999, textvariable=qty_var, width=6)
            change_type = ttk.Combobox(row, textvariable=type_var, values=["增配", "减配"], width=8, state="readonly")
            price_label = ttk.Label(row, text=option.get("price", ""), width=12)
            check.pack(side=LEFT)
            qty.pack(side=LEFT, padx=(0, 8))
            change_type.pack(side=LEFT, padx=(0, 8))
            price_label.pack(side=LEFT, padx=(0, 8))
            for widget in (row, check, qty, change_type, price_label):
                widget.bind("<MouseWheel>", self.option_frame.on_mousewheel)
            if option.get("children"):
                button = ttk.Button(
                    row,
                    text="查看包内容",
                    command=lambda item=option: self.show_option_package_contents(item),
                )
                button.pack(side=RIGHT, padx=(8, 0))
                button.bind("<MouseWheel>", self.option_frame.on_mousewheel)
            label = ttk.Label(row, text=option.get("item_display", ""), wraplength=360)
            label.pack(side=LEFT, fill="x", expand=True, anchor="w")
            label.bind("<MouseWheel>", self.option_frame.on_mousewheel)
        self.update_current_price()

    def show_option_package_contents(self, option):
        children = option.get("children") or []
        dialog = Toplevel(self.root)
        dialog.title(f"{option.get('item_display', '包')} - 包内容")
        dialog.geometry("900x480")
        dialog.minsize(720, 360)
        dialog.transient(self.root)
        try:
            if WINDOW_ICON_FILE.exists():
                dialog.iconbitmap(default=str(WINDOW_ICON_FILE))
        except Exception:
            pass

        wrapper = ttk.Frame(dialog, padding=12)
        wrapper.pack(fill=BOTH, expand=True)
        ttk.Label(
            wrapper,
            text=f"{option.get('item_display', '包')}（共 {len(children)} 项）",
            style="DialogTitle.TLabel",
        ).pack(anchor="w", pady=(0, 10))

        table_frame = ttk.Frame(wrapper)
        table_frame.pack(fill=BOTH, expand=True)
        columns = ("seq", "component", "name", "code", "model_code", "mark")
        tree = ttk.Treeview(table_frame, columns=columns, show="headings", height=15)
        headings = {
            "seq": "序号",
            "component": "部件",
            "name": "名称",
            "code": "编码",
            "model_code": "代号",
            "mark": "配置标记",
        }
        widths = {"seq": 55, "component": 130, "name": 210, "code": 190, "model_code": 170, "mark": 90}
        for column in columns:
            tree.heading(column, text=headings[column])
            tree.column(column, width=widths[column], anchor="w")
        for index, child in enumerate(children, start=1):
            tree.insert(
                "",
                "end",
                values=(
                    index,
                    compact_display_text(child.get("component", "")),
                    child.get("name", ""),
                    child.get("code", ""),
                    code_or_slash(child.get("model_code", "")),
                    child.get("mark", ""),
                ),
            )

        vertical_scroll = ttk.Scrollbar(table_frame, orient=VERTICAL, command=tree.yview)
        horizontal_scroll = ttk.Scrollbar(table_frame, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=vertical_scroll.set, xscrollcommand=horizontal_scroll.set)
        tree.grid(row=0, column=0, sticky="nsew")
        vertical_scroll.grid(row=0, column=1, sticky="ns")
        horizontal_scroll.grid(row=1, column=0, sticky="ew")
        table_frame.rowconfigure(0, weight=1)
        table_frame.columnconfigure(0, weight=1)

        ttk.Button(wrapper, text="关闭", command=dialog.destroy).pack(anchor="e", pady=(10, 0))

    def machine_price_value(self):
        if not self.model_var.get() or not self.form_var.get():
            return 0.0
        try:
            items = basic_config_items(self.db, self.model_var.get(), self.form_var.get())
        except Exception:
            return 0.0
        for item in items:
            price = parse_price_value(item.get("price", ""))
            if price:
                return price
        return 0.0

    def selected_option_price_value(self):
        total = 0.0
        for idx, var in enumerate(self.option_vars):
            if not var.get() or idx >= len(self.config_option_items):
                continue
            item = self.config_option_items[idx]
            qty = parse_price_value(self.option_qty_vars[idx].get() if idx < len(self.option_qty_vars) else "1") or 1
            price = parse_price_value(item.get("price", ""))
            change_type = clean_text(self.option_type_vars[idx].get()) if idx < len(self.option_type_vars) else "增配"
            sign = -1 if "减" in change_type else 1
            total += sign * price * qty
        return total

    def update_current_price(self):
        machine_price = self.machine_price_value()
        option_price = self.selected_option_price_value()
        total_price = machine_price + option_price
        self.machine_price_var.set(format_price_value(machine_price) if machine_price else "")
        self.option_price_var.set(format_price_value(option_price) if option_price else "")
        self.total_price_var.set(format_price_value(total_price) if total_price else "")

    def refresh_basic_config(self):
        for item_id in self.basic_config_tree.get_children():
            self.basic_config_tree.delete(item_id)
        if not self.model_var.get() or not self.form_var.get():
            return
        try:
            items = basic_config_items(self.db, self.model_var.get(), self.form_var.get())
        except Exception:
            items = []
        for idx, item in enumerate(items, start=1):
            self.basic_config_tree.insert(
                "",
                "end",
                values=(
                    idx,
                    compact_display_text(item.get("component", "")),
                    item.get("name", ""),
                    code_or_slash(item.get("model_code", "")),
                    quantity_from_mark(item.get("mark", "")),
                ),
            )
        self.update_current_price()

    def reload_db(self):
        self.db = load_database()
        self.refresh_models()

    def open_research_import_module(self):
        self.open_protected_module("配置及价格导入模块", self.show_research_import_window)

    def open_audit_publish_module(self):
        self.open_protected_module("审核并发布", self.show_audit_publish_window)

    def open_protected_module(self, title, on_success):
        password_window = Toplevel(self.root)
        password_window.title(title)
        password_window.geometry("420x240")
        password_window.transient(self.root)
        password_window.grab_set()
        frame = ttk.Frame(password_window, padding=14)
        frame.pack(fill=BOTH, expand=True)

        ttk.Label(frame, text="工号").grid(row=0, column=0, sticky="w", padx=(0, 8), pady=5)
        operator_var = StringVar(value=self.research_operator_id)
        operator_entry = ttk.Entry(frame, textvariable=operator_var, width=30)
        operator_entry.grid(row=0, column=1, sticky="we", pady=5)

        ttk.Label(frame, text="密码").grid(row=1, column=0, sticky="w", padx=(0, 8), pady=5)
        password_var = StringVar()
        entry = ttk.Entry(frame, textvariable=password_var, show="*", width=28)
        entry.grid(row=1, column=1, sticky="we", pady=5)

        ttk.Label(frame, text="时间").grid(row=2, column=0, sticky="w", padx=(0, 8), pady=5)
        time_var = StringVar(value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        time_entry = ttk.Entry(frame, textvariable=time_var, width=30, state="readonly")
        time_entry.grid(row=2, column=1, sticky="we", pady=5)
        frame.columnconfigure(1, weight=1)
        operator_entry.focus_set()

        def verify():
            if not clean_text(operator_var.get()):
                messagebox.showerror("缺少工号", "请输入工号。", parent=password_window)
                return
            if not clean_text(time_var.get()):
                messagebox.showerror("缺少时间", "请输入时间。", parent=password_window)
                return
            is_admin_login = clean_text(operator_var.get()) == "1" and password_var.get() == "1"
            if password_var.get() != "zlzk.123456789" and not is_admin_login:
                messagebox.showerror("密码错误", f"密码不正确，无法进入{title}。", parent=password_window)
                return
            self.research_operator_id = "00796274" if is_admin_login else clean_text(operator_var.get())
            self.research_operation_time = clean_text(time_var.get())
            password_window.destroy()
            on_success()

        buttons = ttk.Frame(frame)
        buttons.grid(row=3, column=0, columnspan=2, sticky="w", pady=(12, 0))
        ttk.Button(buttons, text="进入", command=verify).pack(side=LEFT)
        ttk.Button(buttons, text="取消", command=password_window.destroy).pack(side=LEFT, padx=(8, 0))
        for widget in (operator_entry, entry):
            widget.bind("<Return>", lambda _e: verify())

    def show_research_import_window(self):
        window = Toplevel(self.root)
        window.title("配置及价格导入模块")
        window.geometry("520x260")
        window.transient(self.root)
        frame = ttk.Frame(window, padding=16)
        frame.pack(fill=BOTH, expand=True)
        ttk.Label(frame, text="配置及价格导入模块", font=("Microsoft YaHei UI", 13, "bold")).pack(anchor="w", pady=(0, 10))
        ttk.Label(frame, text=f"当前工号：{self.research_operator_id}    记录时间：{self.research_operation_time}", foreground="#555").pack(anchor="w", pady=(0, 8))
        ttk.Label(frame, text="以下功能用于维护软件内置数据库，请谨慎操作。", foreground="#555").pack(anchor="w", pady=(0, 12))
        ttk.Button(frame, text="导入增减配价格表", command=self.on_import_price).pack(fill="x", pady=4)
        ttk.Button(frame, text="导入产品安装参数表", command=self.on_import_product_cfg).pack(fill="x", pady=4)
        ttk.Button(frame, text="导入机型配置表", command=self.on_import_tower_config).pack(fill="x", pady=4)

    def record_change(self, content):
        append_change_log(self.research_operator_id, self.research_operation_time, content)

    def on_import_price(self):
        path = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx *.xlsm")])
        if not path:
            return
        try:
            products, options = import_price_file(path)
            self.record_change(f"导入增减配价格表：{Path(path).name}；识别机型 {products} 个，增减配 {options} 条。")
            self.reload_db()
            messagebox.showinfo("导入完成", f"已导入 {products} 个机型、{options} 条增减配。")
        except Exception as exc:
            messagebox.showerror("导入失败", str(exc))

    def on_import_product_cfg(self):
        path = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx *.xlsm")])
        if not path:
            return
        try:
            count = import_product_cfg_file(path)
            self.record_change(f"导入产品安装参数表：{Path(path).name}；导入/更新安装形式参数 {count} 条。")
            self.reload_db()
            messagebox.showinfo("导入完成", f"已导入 {count} 条安装形式/塔高/臂长/容绳量。")
        except Exception as exc:
            messagebox.showerror("导入失败", str(exc))

    def on_import_tower_config(self):
        products = sorted(model for model in self.db.get("products", {}) if is_visible_product_model(model))
        if not products:
            messagebox.showinfo("缺少机型", "请先导入增减配价格表，系统识别到产品型号后再导入机型配置表。")
            return

        dialog = Toplevel(self.root)
        dialog.title("选择配置表所属机型")
        dialog.geometry("420x150")
        dialog.transient(self.root)
        dialog.grab_set()
        frame = ttk.Frame(dialog, padding=14)
        frame.pack(fill=BOTH, expand=True)
        ttk.Label(frame, text="请选择这张配置表对应的产品型号：").pack(anchor="w")
        selected_model = StringVar(value=self.model_var.get() if self.model_var.get() in products else products[0])
        combo = ttk.Combobox(frame, textvariable=selected_model, values=products, state="readonly", width=38)
        combo.pack(fill="x", pady=(8, 12))

        def choose_file():
            model_name = selected_model.get()
            dialog.destroy()
            path = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx *.xlsm")])
            if not path:
                return
            try:
                model, count, old_sources = import_tower_config_file(path, model_name=model_name)
                action = "覆盖旧配置表" if old_sources else "新增配置表"
                old_text = f"；旧文件：{'、'.join(old_sources)}" if old_sources else ""
                self.record_change(f"导入机型配置表：{model}；{action}；新文件：{Path(path).name}；主要部件 {count} 条；发布状态：待发布{old_text}。")
                self.reload_db()
                messagebox.showinfo("导入完成", f"已将配置表导入到 {model}，共 {count} 条主要部件配置。{action}，当前发布状态为待发布。")
            except Exception as exc:
                messagebox.showerror("导入失败", str(exc))

        buttons = ttk.Frame(frame)
        buttons.pack(fill="x")
        ttk.Button(buttons, text="下一步：选择配置表文件", command=choose_file).pack(side=LEFT)
        ttk.Button(buttons, text="取消", command=dialog.destroy).pack(side=LEFT, padx=(8, 0))

    def on_delete_tower_config(self):
        configs = {}
        for cfg in list(self.db.get("components_by_model", {}).values()) + list(self.db.get("config_exports_by_model", {}).values()):
            model = clean_text(cfg.get("model")) or clean_text(cfg.get("source_file"))
            if model:
                configs[normalize_key(model)] = model
        models = sorted(configs.values())
        if not models:
            messagebox.showinfo("没有配置表", "当前数据库里没有已导入的机型配置表。")
            return

        dialog = Toplevel(self.root)
        dialog.title("删除机型配置表")
        dialog.geometry("440x170")
        dialog.transient(self.root)
        dialog.grab_set()
        frame = ttk.Frame(dialog, padding=14)
        frame.pack(fill=BOTH, expand=True)
        ttk.Label(frame, text="请选择要删除配置表的产品型号：").pack(anchor="w")
        selected_model = StringVar(value=self.model_var.get() if self.model_var.get() in models else models[0])
        combo = ttk.Combobox(frame, textvariable=selected_model, values=models, state="readonly", width=40)
        combo.pack(fill="x", pady=(8, 10))
        ttk.Label(frame, text="只删除该机型配置表，不删除产品型号、安装参数和增减配价格数据。", foreground="#555").pack(anchor="w")

        def delete_selected():
            model_name = selected_model.get()
            if not messagebox.askyesno("确认删除", f"确认删除 {model_name} 的配置表吗？", parent=dialog):
                return
            try:
                sources, forms = delete_tower_config(model_name)
                if not sources:
                    messagebox.showinfo("未删除", f"{model_name} 没有找到可删除的配置表。", parent=dialog)
                    return
                self.record_change(
                    f"删除机型配置表：{model_name}；删除文件：{'、'.join(sources)}；安装形式：{'、'.join(forms) if forms else '无'}。"
                )
                dialog.destroy()
                self.reload_db()
                messagebox.showinfo("删除完成", f"已删除 {model_name} 的配置表。")
            except Exception as exc:
                messagebox.showerror("删除失败", str(exc), parent=dialog)

        buttons = ttk.Frame(frame)
        buttons.pack(fill="x", pady=(12, 0))
        ttk.Button(buttons, text="删除", command=delete_selected).pack(side=LEFT)
        ttk.Button(buttons, text="取消", command=dialog.destroy).pack(side=LEFT, padx=(8, 0))

    def show_audit_publish_window(self):
        self.db = load_database()
        products = {
            model: product
            for model, product in self.db.get("products", {}).items()
            if is_visible_product_model(model)
        }
        if not products:
            messagebox.showinfo("审核并发布", "请先导入增减配价格表，系统识别到产品型号后再审核发布。")
            return

        window = Toplevel(self.root)
        window.title("审核并发布")
        window.geometry("1060x500")
        wrapper = ttk.Frame(window, padding=10)
        wrapper.pack(fill=BOTH, expand=True)

        summary_var = StringVar()
        ttk.Label(wrapper, textvariable=summary_var, font=("Microsoft YaHei UI", 10, "bold")).pack(anchor="w")
        ttk.Label(
            wrapper,
            text="说明：导入或覆盖机型配置表后，该机型状态会变为待发布；审核确认后点击发布。",
            foreground="#555",
        ).pack(anchor="w", pady=(0, 8))

        columns = ("model", "code", "config_status", "publish_status", "source", "forms")
        tree = ttk.Treeview(wrapper, columns=columns, show="headings", height=14)
        headings = {
            "model": "产品型号",
            "code": "产品编码",
            "config_status": "配置表状态",
            "publish_status": "发布状态",
            "source": "导入文件",
            "forms": "匹配安装形式",
        }
        widths = {
            "model": 170,
            "code": 100,
            "config_status": 100,
            "publish_status": 100,
            "source": 260,
            "forms": 260,
        }
        for col in columns:
            tree.heading(col, text=headings[col])
            tree.column(col, width=widths[col], anchor="w")

        table_frame = ttk.Frame(wrapper)
        table_frame.pack(fill=BOTH, expand=True)
        scrollbar = ttk.Scrollbar(table_frame, orient=VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        tree.pack(side=LEFT, fill=BOTH, expand=True)
        scrollbar.pack(side=RIGHT, fill="y")

        def populate():
            self.db = load_database()
            tree.delete(*tree.get_children())
            configs = self.db.get("components_by_model", {})
            cfg_by_key = {cfg.get("model_key"): cfg for cfg in configs.values()}
            imported_count = 0
            published_count = 0
            pending_count = 0
            def status_sort_key(model_name):
                cfg = cfg_by_key.get(normalize_key(model_name))
                if cfg and not is_config_published(self.db, model_name) and not is_config_rejected(self.db, model_name):
                    return (0, model_name)
                if cfg and is_config_rejected(self.db, model_name):
                    return (1, model_name)
                if cfg:
                    return (2, model_name)
                return (3, model_name)

            for model_name in sorted(products, key=status_sort_key):
                product = products[model_name]
                cfg = cfg_by_key.get(normalize_key(model_name))
                forms = self.db.get("forms_by_code", {}).get(product.get("product_code", ""), [])
                if cfg:
                    imported_count += 1
                    cfg_forms = cfg.get("forms", {})
                    matched_forms = [
                        form.get("install_form", "")
                        for form in forms
                        if normalize_form(form.get("install_form", "")) in cfg_forms
                    ]
                    config_status = "已导入"
                    if is_config_published(self.db, model_name):
                        publish_status = "已发布"
                    elif is_config_rejected(self.db, model_name):
                        publish_status = "已拒绝"
                    else:
                        publish_status = "待发布"
                    if publish_status == "已发布":
                        published_count += 1
                    elif publish_status == "待发布":
                        pending_count += 1
                    source = cfg.get("source_file", "")
                    forms_text = "、".join(matched_forms) if matched_forms else "无匹配安装形式"
                else:
                    config_status = "未导入"
                    publish_status = "未发布"
                    source = ""
                    forms_text = ""
                tree.insert(
                    "",
                    "end",
                    values=(
                        model_name,
                        product.get("product_code", ""),
                        config_status,
                        publish_status,
                        source,
                        forms_text,
                    ),
                )
            summary_var.set(
                f"审核并发布：共 {len(products)} 个机型，已导入 {imported_count} 个，待发布 {pending_count} 个，已发布 {published_count} 个。"
            )

        def publish_selected():
            selection = tree.selection()
            if not selection:
                messagebox.showinfo("未选择机型", "请先选择要发布的机型。", parent=window)
                return
            values = tree.item(selection[0], "values")
            model_name = values[0]
            config_status = values[2]
            publish_status = values[3]
            source = values[4]
            if config_status != "已导入":
                messagebox.showwarning("不能发布", f"{model_name} 尚未导入配置表。", parent=window)
                return
            if publish_status == "已发布":
                messagebox.showinfo("已发布", f"{model_name} 已经是发布状态。", parent=window)
                return
            if publish_status == "已拒绝":
                messagebox.showwarning("不能发布", f"{model_name} 已被拒绝发布，请重新导入配置表后再审核发布。", parent=window)
                return
            try:
                publish_tower_config(model_name)
                self.record_change(f"发布机型配置表：{model_name}；导入文件：{source}。")
                populate()
                self.reload_db()
                messagebox.showinfo("发布完成", f"{model_name} 已发布。", parent=window)
            except Exception as exc:
                messagebox.showerror("发布失败", str(exc), parent=window)

        def reject_selected():
            selection = tree.selection()
            if not selection:
                messagebox.showinfo("未选择机型", "请先选择要拒绝发布的机型。", parent=window)
                return
            values = tree.item(selection[0], "values")
            model_name = values[0]
            config_status = values[2]
            publish_status = values[3]
            source = values[4]
            if config_status != "已导入":
                messagebox.showwarning("不能拒绝", f"{model_name} 尚未导入配置表。", parent=window)
                return
            if publish_status == "已发布":
                messagebox.showwarning("不能拒绝", f"{model_name} 已发布，如需更改请重新导入配置表覆盖后再审核。", parent=window)
                return
            if publish_status == "已拒绝":
                messagebox.showinfo("已拒绝", f"{model_name} 已经是拒绝状态。", parent=window)
                return
            if not messagebox.askyesno("确认拒绝发布", f"确认拒绝发布 {model_name} 吗？拒绝后需要重新导入配置表才能再次审核发布。", parent=window):
                return
            try:
                reject_tower_config(model_name)
                self.record_change(f"拒绝发布机型配置表：{model_name}；导入文件：{source}；需重新导入配置表后再审核发布。")
                populate()
                self.reload_db()
                messagebox.showinfo("已拒绝发布", f"{model_name} 已拒绝发布，请重新导入配置表后再审核。", parent=window)
            except Exception as exc:
                messagebox.showerror("拒绝失败", str(exc), parent=window)

        button_bar = ttk.Frame(wrapper)
        button_bar.pack(fill="x", pady=(10, 0))
        ttk.Button(button_bar, text="发布选中机型", command=publish_selected).pack(side=LEFT)
        ttk.Button(button_bar, text="拒绝发布", command=reject_selected).pack(side=LEFT, padx=(8, 0))
        ttk.Button(button_bar, text="关闭", command=window.destroy).pack(side=LEFT, padx=(8, 0))
        populate()

    def show_change_log(self):
        self.db = load_database()
        logs = self.db.get("change_log", [])
        window = Toplevel(self.root)
        window.title("后台数据变更记录")
        window.geometry("980x460")
        wrapper = ttk.Frame(window, padding=10)
        wrapper.pack(fill=BOTH, expand=True)
        ttk.Label(
            wrapper,
            text=f"后台数据变更记录：共 {len(logs)} 条。",
            font=("Microsoft YaHei UI", 10, "bold"),
        ).pack(anchor="w", pady=(0, 8))

        columns = ("time", "operator", "content")
        tree = ttk.Treeview(wrapper, columns=columns, show="headings", height=16)
        tree.heading("time", text="时间")
        tree.heading("operator", text="工号")
        tree.heading("content", text="更改内容")
        tree.column("time", width=170, anchor="w")
        tree.column("operator", width=120, anchor="w")
        tree.column("content", width=650, anchor="w")

        for item in reversed(logs):
            tree.insert(
                "",
                "end",
                values=(
                    item.get("time", ""),
                    item.get("operator_id", ""),
                    item.get("content", ""),
                ),
            )

        scrollbar = ttk.Scrollbar(wrapper, orient=VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        tree.pack(side=LEFT, fill=BOTH, expand=True)
        scrollbar.pack(side=RIGHT, fill="y")

    def term_value(self, key):
        widget = self.term_widgets.get(key)
        if widget is not None and widget.winfo_exists():
            return widget.get("1.0", "end").strip()
        return self.term_values.get(key, self.term_defaults.get(key, ""))

    def save_current_user_settings(self):
        save_user_settings(
            {
                "header_address": self.header_address.get(),
                "customer_name": self.customer_name.get(),
                "quote_company": self.quote_company.get(),
                "quote_person": self.quote_person.get(),
                "quote_phone": self.quote_phone.get(),
                "quote_email": self.quote_email.get(),
            }
        )

    def build_quote(self):
        self.sync_quote_dialog_values()
        selected_items = []
        for idx, var in enumerate(self.option_vars):
            if not var.get() or idx >= len(self.config_option_items):
                continue
            item = dict(self.config_option_items[idx])
            item["quantity"] = clean_text(self.option_qty_vars[idx].get()) or "1"
            item["change_type"] = self.option_type_vars[idx].get() if idx < len(self.option_type_vars) else "增配"
            selected_items.append(item)
        return QuoteInput(
            language=LANG_OPTIONS.get(self.language_label.get(), "zh"),
            model_name=self.model_var.get(),
            install_form=self.form_var.get(),
            quote_date=self.date_var.get(),
            trade_term=self.trade_term.get(),
            trade_place=self.trade_place.get(),
            price=self.price.get(),
            currency=self.currency.get(),
            header_address=self.header_address.get(),
            customer_name=self.customer_name.get(),
            quote_company=self.quote_company.get(),
            quote_person=self.quote_person.get(),
            quote_phone=self.quote_phone.get(),
            quote_email=self.quote_email.get(),
            payment=self.term_value("payment"),
            delivery=self.term_value("delivery"),
            validity=self.term_value("validity"),
            transportation=self.term_value("transportation"),
            warranty=self.term_value("warranty"),
            others=self.term_value("others"),
            selected_option_indexes=[idx for idx, var in enumerate(self.option_vars) if var.get()],
            selected_option_quantities={
                str(idx): clean_text(self.option_qty_vars[idx].get()) or "1"
                for idx, var in enumerate(self.option_vars)
                if var.get()
            },
            selected_option_items=selected_items,
        )

    def on_generate(self):
        if not self.model_var.get() or not self.form_var.get():
            messagebox.showwarning("缺少数据", "请先导入数据并选择产品型号和安装形式。")
            return
        quote = self.build_quote()
        self.save_current_user_settings()
        safe_model = re.sub(r"[^0-9A-Za-z._()-]+", "_", quote.model_name)
        generated_at = datetime.now().strftime("%Y-%m-%d_%H%M")
        quote_folder = available_output_folder(OUTPUT_DIR / f"塔机报价_{safe_model}_{generated_at}")
        output = quote_folder / f"中联塔机报价单_{safe_model}_{generated_at}.pdf"
        try:
            make_pdf(self.db, quote, output)
            generated = [output]
            if quote.selected_option_items:
                ltc_output = available_output_path(quote_folder / f"LTC选配指导文件_{safe_model}_{generated_at}.pdf")
                make_ltc_option_pdf(quote, ltc_output)
                generated.append(ltc_output)
            messagebox.showinfo("生成完成", "已生成:\n" + "\n".join(str(path) for path in generated))
            open_folder(quote_folder)
        except Exception as exc:
            messagebox.showerror("生成失败", str(exc))

    def on_generate_config_list(self, list_type):
        if not self.model_var.get() or not self.form_var.get():
            messagebox.showwarning("缺少数据", "请先选择产品型号和安装形式。")
            return
        title = "基本配置清单" if list_type == "basic" else "增减配清单"
        safe_model = re.sub(r"[^0-9A-Za-z._()-]+", "_", self.model_var.get())
        safe_form = re.sub(r"[^0-9A-Za-z._()-]+", "_", self.form_var.get())
        output = categorized_output_path("配置清单", f"{title}_{safe_model}_{safe_form}_{date.today().isoformat()}.xlsx")
        try:
            output = make_config_list_excel(
                self.db,
                self.model_var.get(),
                self.form_var.get(),
                list_type,
                output,
                lang=LANG_OPTIONS.get(self.language_label.get(), "zh"),
            )
            self.db = load_database()
            messagebox.showinfo("生成完成", f"已生成:\n{output}")
            open_containing_folder(output)
        except Exception as exc:
            messagebox.showerror(
                "生成失败",
                f"机型：{self.model_var.get()}\n安装形式：{self.form_var.get()}\n清单类型：{title}\n错误：{exc}",
            )

    def on_generate_combined_config_list(self):
        if not self.model_var.get():
            messagebox.showwarning("缺少数据", "请先选择产品型号。")
            return
        safe_model = safe_filename_stem(self.model_var.get())
        version_suffix = workbook_version_filename_suffix(self.db, self.model_var.get())
        suffix_text = f"_{version_suffix}" if version_suffix else ""
        output = categorized_output_path("配置及增减配清单", f"{safe_model}配置及增减配清单{suffix_text}.xlsx")
        try:
            output = make_combined_config_option_excel(
                self.db,
                self.model_var.get(),
                output,
                lang=LANG_OPTIONS.get(self.language_label.get(), "zh"),
            )
            messagebox.showinfo("生成完成", f"已生成:\n{output}")
            open_containing_folder(output)
        except Exception as exc:
            messagebox.showerror(
                "生成失败",
                f"机型：{self.model_var.get()}\n清单类型：配置及增减配清单\n错误：{exc}",
            )

    def on_close(self):
        self.sync_quote_dialog_values()
        self.save_current_user_settings()
        self.root.destroy()


def main():
    ensure_dirs()
    root = Tk()
    QuotationApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
