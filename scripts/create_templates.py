from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
TEMPLATE_DIR = ROOT / "templates"


def style_sheet(ws):
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 22
    ws.freeze_panes = "A2"


def build_models():
    wb = Workbook()
    ws = wb.active
    ws.title = "models"
    headers = [
        "model",
        "base_form",
        "type_zh",
        "type_en",
        "internal_climbing_height",
        "jib_length",
        "max_load",
        "tip_load",
        "mast_section_dimension",
        "status_zh",
        "status_en",
        "price_terms",
        "currency",
        "unit_price",
        "overview_zh",
        "overview_en",
        "transportation",
    ]
    ws.append(headers)
    ws.append(
        [
            "T2850-120V",
            "内爬式",
            "动臂式",
            "Luffing",
            "32.5m",
            "40m",
            "32t",
            "19.6t",
            "2.5*2.5*5.98m",
            "内爬式",
            "Internal Climbing",
            "FOB 上海港",
            "CNY",
            "4,850,000",
            "1套起升变频机构\n2套回转变频机构\n1套变幅变频机构\n1套平衡臂（含平衡配重）\n1套40m起重臂\n1套上下回转平台\n1套内爬装置（不含液压油）\n1套吊钩\n1个驾驶室（含空调）\n全套电气控制系统\n全套电缆、钢丝绳和标准附件\n全套标准技术文件和备件图册",
            "01 set of hoisting mechanism with inverter\n02 sets of slewing mechanism with inverter\n01 set of luffing mechanism with inverter\n01 set of balance boom with ballast\n01 set of 40m jib\n01 set of slewing bracket\n01 set of internal climbing system, hydraulic oil excluded\n01 set of hook\n01 unit of cabin with air conditioner\nComplete electric control system\nComplete electric cable, wire rope and standard accessories\nComplete technical documents and parts catalog",
            "11*40'HQ + 480m3",
        ]
    )
    style_sheet(ws)
    ws.column_dimensions["O"].width = 48
    ws.column_dimensions["P"].width = 48
    return wb


def build_options():
    wb = Workbook()
    ws = wb.active
    ws.title = "options"
    headers = ["model", "base_form", "action", "item_zh", "item_en", "price_delta", "remark_zh", "remark_en"]
    ws.append(headers)
    ws.append(["T2850-120V", "内爬式", "add", "增加10m起重臂", "Add 10m jib", "+ CNY 180,000", "含相应拉杆和销轴", "Including related tie bars and pins"])
    ws.append(["T2850-120V", "内爬式", "add", "远程监控系统", "Remote monitoring system", "+ CNY 65,000", "含一年平台服务", "Including one-year platform service"])
    ws.append(["T2850-120V", "内爬式", "deduct", "取消驾驶室空调", "Remove cabin air conditioner", "- CNY 8,000", "", ""])
    style_sheet(ws)
    return wb


def main():
    TEMPLATE_DIR.mkdir(parents=True, exist_ok=True)
    build_models().save(TEMPLATE_DIR / "models_template.xlsx")
    build_options().save(TEMPLATE_DIR / "options_template.xlsx")


if __name__ == "__main__":
    main()
